import yaml
import os
import shutil
import openpyxl
from datetime import datetime
from core_engine.data_reader import DataReader
from core_engine.transformers import apply_formula
from plugins.aviation_plugin import AviationPlugin
from plugins.macro_plugin import MacroPlugin
from plugins.electronics_plugin import ElectronicsPlugin
from plugins.baijiu_plugin import BaijiuPlugin
from plugins.metal_plugin import MetalPlugin
from plugins.chemical_plugin import ChemicalPlugin
from plugins.pesticide_plugin import PesticidePlugin
from plugins.macro_plugin import MacroPlugin
from plugins.public_plugin import PublicPlugin
from plugins.bank_plugin import BankPlugin
from plugins.medical_plugin import MedicalPlugin
from plugins.media_internet_plugin import MediaInternetPlugin
from plugins.appliance_plugin import AppliancePlugin
from plugins.petrochemical_plugin import PetrochemicalPlugin
from plugins.electronics_plugin import ElectronicsPlugin
from plugins.baijiu_plugin import BaijiuPlugin
from plugins.farm_plugin import FarmPlugin
from plugins.farm_export_plugin import FarmPlugin as FarmExportPlugin
from plugins.farm_zlzd_plugin import FarmPlugin as FarmZlzdPlugin
from plugins.farm_chick_plugin import FarmPlugin as FarmChickPlugin
from plugins.building_plugin import BuildingPlugin
from plugins.decoration_plugin import DecorationPlugin
from plugins.PO_plugin import PoPlugin

# === Monkey-patch: 修复 openpyxl 无法读取含 ExternalData 图表的 bug ===
# ExternalData.id 定义为 String() 但不支持命名空间属性 r:id，
# 导致 id 解析为 None 时抛出 TypeError，图表被静默丢弃。
# 修复: 重新创建 String 描述符，启用 allow_none 并设置 REL_NS 命名空间映射。
# 详见: openpyxl/chart/chartspace.py 中 `id = String()  # Needs namespace`
from openpyxl.chart.chartspace import ExternalData
from openpyxl.descriptors.base import String
from openpyxl.xml.constants import REL_NS

_new_id = String(allow_none=True, namespace=REL_NS)
_new_id.name = "id"  # 描述符的 name 属性必须设置，否则无法读写实例属性
ExternalData.id = _new_id

# 同步更新 __namespaced__，使 parse/serialize 时能正确处理 r:id 属性
_ns_id = "{%s}%s" % (REL_NS, "id")
_namespaced = [(k, ns) for k, ns in ExternalData.__namespaced__ if k != "id"]
_namespaced.append(("id", _ns_id))
ExternalData.__namespaced__ = tuple(_namespaced)
# ==================================================================


class PipelineEngine:
    def __init__(self, config_path):
        with open(config_path, 'r', encoding='utf-8') as f:
            self.config = yaml.safe_load(f)
            
        self.source_file = self.config['source_file']
        self.target_file = self.config['target_file']
        self.output_dir = self.config.get('output_dir', './output')
        
        # 动作注册表
        self.actions_registry = {
            'apply_formula': apply_formula,
            # 航空动作
            'aviation_write_airline_sheet': AviationPlugin.aviation_write_airline_sheet,
            'aviation_apply_yoy_formulas': AviationPlugin.aviation_apply_yoy_formulas,
            'aviation_apply_yoy_diff_formulas': AviationPlugin.aviation_apply_yoy_diff_formulas,
            'aviation_apply_yoy19_formulas': AviationPlugin.aviation_apply_yoy19_formulas,
            'aviation_apply_diff19_formulas': AviationPlugin.aviation_apply_diff19_formulas,
            'aviation_clear_early_years_data': AviationPlugin.aviation_clear_early_years_data,
            'aviation_adjust_format_after_full_year': AviationPlugin.aviation_adjust_format_after_full_year,
            'aviation_clear_ytd_2018_diff_data': AviationPlugin.aviation_clear_ytd_2018_diff_data,
            'aviation_clear_ax_ay_az_columns': AviationPlugin.aviation_clear_ax_ay_az_columns,
            'aviation_write_monthly_report_header_info': AviationPlugin.aviation_write_monthly_report_header_info,
            'aviation_write_quarterly_report_header_info': AviationPlugin.aviation_write_quarterly_report_header_info,
            'aviation_write_report_data': AviationPlugin.aviation_write_report_data,
            # 宏观动作
            'write_indicator_group': MacroPlugin.macro_write_indicator_group,
            'compute_indicator_formula': MacroPlugin.macro_compute_indicator_formula,
            'create_pivot_table': MacroPlugin.macro_create_pivot_table,
            'macro_create_festival_pivot': MacroPlugin.macro_create_festival_pivot,
            'macro_create_chuxi_pivot': MacroPlugin.macro_create_chuxi_pivot,
            'macro_create_weekly_pivot': MacroPlugin.macro_create_weekly_pivot,
            'macro_create_yearly_date_scaffold': MacroPlugin.macro_create_yearly_date_scaffold,
            'calc_indicator_yoy_mom': MacroPlugin.macro_calc_indicator_yoy_mom, 
            # 电子动作
            'electronics_write_sheet': ElectronicsPlugin.electronics_write_sheet,
            'electronics_update_chart_ranges': ElectronicsPlugin.electronics_update_chart_ranges,
            # 白酒动作
            'baijiu_write_sheet': BaijiuPlugin.baijiu_write_sheet,
            'baijiu_write_sheet_by_code': BaijiuPlugin.baijiu_write_sheet_by_code,
            'baijiu_fill_summary_rows': BaijiuPlugin.baijiu_fill_summary_rows,
            'baijiu_fill_zuotu': BaijiuPlugin.baijiu_fill_zuotu,
            'baijiu_finalize_charts': BaijiuPlugin.baijiu_finalize_charts,
            # 有色金属动作
            'metal_write_sheet': MetalPlugin.metal_write_sheet,
            # 化工动作
            'chemical_write_sheet': ChemicalPlugin.chemical_write_sheet,
            # 农药动作
            'pesticide_write_sheet': PesticidePlugin.pesticide_write_sheet,
            'pesticide_write_summary': PesticidePlugin.pesticide_write_summary,
            # 公用事业动作
            'public_write_data': PublicPlugin.public_write_data,
            'public_elec_write_title': PublicPlugin.public_elec_write_title,
            'public_t3_write_header': PublicPlugin.public_t3_write_header,
            # 银行动作
            'bank_write_data': BankPlugin.bank_write_data,

            'bank_commercial_write_data': BankPlugin.bank_commercial_write_data,
            'bank_commercial_formula': BankPlugin.bank_commercial_formula,
            # 医疗健康动作
            'medical_write_data': MedicalPlugin.medical_write_data,
            'medical_write_quarter': MedicalPlugin.medical_write_quarter,
            'medical_write_formula': MedicalPlugin.medical_write_formula,
            'medical_merge_quarter_data': MedicalPlugin.medical_merge_quarter_data,
            # 农业动作
            'farm_write_data': FarmExportPlugin.farm_write_data,
            'farm_zlzd_write_data': FarmZlzdPlugin.farm_write_data,
            'farm_zlzd_compute_period_avg': FarmZlzdPlugin.farm_compute_period_avg,
            # 鸡数据库
            'farm_chick_write_data': FarmChickPlugin.farm_write_data,
            'farm_write_linechart': FarmExportPlugin.farm_write_linechart,
            'farm_export_create_pivot': FarmExportPlugin.macro_create_pivot_table,

            # 装饰动作
            'decoration_write_sheet': DecorationPlugin.decoration_write_sheet,
            'decoration_write_formula': DecorationPlugin._render_template,

            # PO/建筑装饰动作
            'PO_write_data': PoPlugin.PO_write_data,
            'PO_write_formula': PoPlugin.PO_write_formula,

            # 建筑动作
            'building_write_data': BuildingPlugin.building_write_data,
            'building_write_formula': BuildingPlugin.building_write_formula,
            'building_write_formula_with_year_lookup': BuildingPlugin.building_write_formula_with_year_lookup,
            'building_calc_monthly_from_cumulative': BuildingPlugin.building_calc_monthly_from_cumulative,
            'building_update_latest_value': BuildingPlugin.update_latest_value,

            # 互联网传媒动作
            'media_internet_write_sheet': MediaInternetPlugin.media_internet_write_sheet,
            'media_internet_apply_yoy': MediaInternetPlugin.media_internet_apply_yoy,
            # 家电动作
            'appliance_write_sheet': AppliancePlugin.appliance_write_sheet,
            # 石油化工动作
            'petrochemical_write_sheet': PetrochemicalPlugin.petrochemical_write_sheet,
            'petrochemical_apply_weekly_formulas': PetrochemicalPlugin.petrochemical_apply_weekly_formulas,
            'petrochemical_apply_monthly_formulas': PetrochemicalPlugin.petrochemical_apply_monthly_formulas,
            'petrochemical_apply_yearly_formulas': PetrochemicalPlugin.petrochemical_apply_yearly_formulas,
        }
        
    def _create_backup(self):
        os.makedirs(self.output_dir, exist_ok=True)
        file_name = os.path.basename(self.target_file)
        name, ext = os.path.splitext(file_name)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file = os.path.join(self.output_dir, f"{name}_{timestamp}{ext}")
        shutil.copy2(self.target_file, backup_file)
        print(f"[{self.config['industry']}] 已创建目标文件副本: {backup_file}")
        return backup_file
        
    def run(self):
        print(f"\n=============================================")
        print(f">> 开始执行 Pipeline: {self.config['industry']}")
        print(f"=============================================\n")
        
        # 1. 预加载所有数据到内存 (DataReader自带缓存机制)
        reader = DataReader(self.source_file)
        
        # 2. 创建输出副本并在内存中打开
        backup_file = self._create_backup()
        have_vba = backup_file.lower().endswith('.xlsm')
        load_kwargs = {}
        if have_vba:
            load_kwargs['keep_vba'] = True
        wb = openpyxl.load_workbook(backup_file, **load_kwargs)
        
        # 3. 遍历并执行每个 sheet 的任务
        defaults = self.config.get('defaults', {})
        for sheet_config in self.config['sheets']:
            sheet_name = sheet_config['sheet_name']
            print(f"\n>> 开始处理工作表: [{sheet_name}]")
            
            if sheet_name not in wb.sheetnames:
                print(f"!! 警告：模板中不存在工作表 {sheet_name}，跳过。")
                print(f"   模板中可用的工作表: {wb.sheetnames}")
                continue
                
            ws = wb[sheet_name]
            
            # 构建上下文环境
            context = {
                'wb': wb,
                'ws': ws,
                'data_reader': reader,
                'sheet_config': sheet_config,
                'defaults': defaults
            }
            
            # 执行配置中的 actions
            for action_cfg in sheet_config.get('actions', []):
                action_type = action_cfg['type']
                if action_type in self.actions_registry:
                    # 混合默认参数
                    params = {**defaults, **action_cfg}
                    self.actions_registry[action_type](context, params)
                else:
                        print(f"!! 未知动作: {action_type}")
                    
            # 执行后处理 post_processes
            for post_cfg in sheet_config.get('post_processes', []):
                action_type = post_cfg['action']
                if action_type in self.actions_registry:
                    params = post_cfg.get('params', {})
                    self.actions_registry[action_type](context, params)
                else:
                    print(f"⚠️ 未知后处理动作: {action_type}")
                    
        # 4. 一次性保存结果
        print(f"\n[Saving] 正在保存文件...")
        wb.save(backup_file)

        # 5. 执行 finalize_actions（在 save 之后，用于修复 openpyxl 的序列化问题）
        for finalize_cfg in self.config.get('finalize_actions', []):
            action_type = finalize_cfg['action']
            if action_type in self.actions_registry:
                fctx = {'filepath': backup_file, 'config': self.config}
                self.actions_registry[action_type](fctx, finalize_cfg.get('params', {}))
            else:
                print(f"⚠️ 未知 finalize 动作: {action_type}")

        print(f"[OK] Pipeline 执行完成！文件已保存至: {backup_file}\n")
