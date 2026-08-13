import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string as column_letter_to_number, get_column_letter as column_number_to_letter
from datetime import datetime

from core_engine.data_processor import DataProcessor

class FarmPlugin:

    @staticmethod
    def farm_write_data(context, params):
        """
        处理农业基础指标数据，写入到目标工作表中。
        支持 date_col 参数指定日期写入列。
        """
        ws = context['ws']
        reader = context['data_reader']
        sheet_config = context['sheet_config']
        source_sheet = sheet_config['source_sheet']
        indicator_col_map = params.get('indicators') or {}

        if isinstance(indicator_col_map, str):
            print(f"    - 错误：indicators 被解析为字符串 '{indicator_col_map}'，"
                  f"请检查 YAML 配置中冒号后是否有空格")
            return
        if not isinstance(indicator_col_map, dict) or not indicator_col_map:
            print(f"    - 错误：indicators 无效（类型: {type(indicator_col_map).__name__}），跳过写入")
            return

        start_row = params.get('start_row', 11)
        start_date = params.get('start_date', '2014-01-01')
        date_format = params.get('date_format', 'yyyy-mm-dd')
        start_date_ts = pd.to_datetime(start_date)

        date_col_raw = params.get('date_col', 'A')
        try:
            date_col_num = column_letter_to_number(date_col_raw) if isinstance(date_col_raw, str) else int(date_col_raw)
        except Exception:
            date_col_num = 1

        # 1. 读取指标数据
        indicator_dfs = {}
        for indicator in indicator_col_map:
            ind_data = reader.read_indicator_data(source_sheet, indicator)
            df = ind_data["data"].copy()

            if df.empty or '日期' not in df.columns:
                indicator_dfs[indicator] = pd.DataFrame(columns=['日期'])
                print(f"    - 工作表[{source_sheet}] 未找到指标 {indicator}，该列跳过写入")
                continue

            df['日期'] = pd.to_datetime(df['日期'], errors='coerce')
            df = df[df['日期'].notna() & (df['日期'] >= start_date_ts)].copy()
            df = df.sort_values('日期', ascending=True).reset_index(drop=True)

            indicator_dfs[indicator] = df

        # 2. 基础日期表
        first_indicator = list(indicator_col_map.keys())[0]
        first_df = indicator_dfs[first_indicator].copy()

        if first_df.empty:
            print(f"    - {source_sheet} 的第一个指标 {first_indicator} 无有效数据，跳过写入")
            return

        base_df = first_df.copy()
        last_data_row = start_row + len(base_df) - 1

        # 0. 取消合并单元格
        col_nums = set()
        for col in indicator_col_map.values():
            try:
                col_nums.add(column_letter_to_number(col) if isinstance(col, str) else int(col))
            except Exception:
                pass
        all_cols = {date_col_num} | col_nums
        for merged_range in list(ws.merged_cells.ranges):
            if (merged_range.min_row >= start_row or merged_range.max_row >= start_row) and \
               any(merged_range.min_col <= c <= merged_range.max_col for c in all_cols):
                ws.unmerge_cells(str(merged_range))

        # 3. 写入日期列
        for i, row in base_df.iterrows():
            current_row = start_row + i
            current_date = row['日期']
            date_cell = ws.cell(row=current_row, column=date_col_num, value=current_date)
            date_cell.number_format = date_format

        # 4. 写入指标数据
        base_date_list = base_df['日期'].tolist()

        for indicator, col in list(indicator_col_map.items()):
            df = indicator_dfs[indicator].copy()
            value_col = df.columns[-1]
            value_map = dict(zip(df['日期'], df[value_col]))

            try:
                col_num = column_letter_to_number(col) if isinstance(col, str) else int(col)
                if col_num < 1:
                    print(f"    - 跳过列 {col}，计算出的列号 {col_num} 无效")
                    continue
            except Exception as e:
                print(f"    - 解析列号失败，跳过列 {col}: {e}")
                continue

            for i, date in enumerate(base_date_list):
                value = value_map.get(date, None)
                ws.cell(row=start_row + i, column=col_num, value=value)

        print(f"    - 完成农业基础数据写入")

        # 5. 处理公式
        formulas = params.get('formulas', [])
        if formulas:
            for fm_cfg in formulas:
                col_letter = fm_cfg.get('col', '')
                template = fm_cfg.get('template', '')
                fm_start_row = fm_cfg.get('start_row', start_row + 1)
                if not col_letter or not template:
                    continue
                try:
                    col_num = column_letter_to_number(col_letter)
                except Exception:
                    print(f"    - 公式列号解析失败: {col_letter}，跳过")
                    continue
                for r in range(fm_start_row, last_data_row + 1):
                    expr = template.replace('{row}', str(r)).replace('{prev_row}', str(r - 1))
                    ws.cell(row=r, column=col_num, value=expr)
                print(f"    - 公式写入完成: 列 {col_letter}，行 {fm_start_row}~{last_data_row}")

        # 6. 清除残留
        clear_start = last_data_row + 1
        clear_end = 2000

        for r in range(clear_start, clear_end + 1):
            ws.cell(row=r, column=date_col_num, value=None)

        for col in indicator_col_map.values():
            try:
                col_num = column_letter_to_number(col) if isinstance(col, str) else int(col)
                for r in range(clear_start, clear_end + 1):
                    ws.cell(row=r, column=col_num, value=None)
            except Exception:
                pass

        print(f"    - 已清除第 {clear_start}-{clear_end} 行残留数据")
