import pandas as pd
from openpyxl.utils import column_index_from_string as col_letter_to_num
from openpyxl.utils import get_column_letter


class MediaInternetPlugin:
    @staticmethod
    def media_internet_write_sheet(context, params):
        """将源数据按 AA 代码读取后写入目标工作表，每个 section 有独立的日期列。"""
        ws = context['ws']
        reader = context['data_reader']
        sheet_config = context['sheet_config']
        source_sheet = sheet_config['source_sheet']
        data_start_row = sheet_config.get('data_start_row', 11)

        sections = sheet_config.get('sections', [])

        for section in sections:
            date_col = section['date_col']
            date_col_num = col_letter_to_num(date_col)
            indicators = section['indicators']

            if isinstance(indicators, dict):
                indicators = list(indicators.items())

            for aa_code, target_col in indicators:
                ind_data = reader.read_indicator_data(source_sheet, aa_code)
                df = ind_data["data"]
                if df.empty:
                    continue

                value_col = df.columns[-1]
                df = df.dropna(subset=['日期']).copy()
                df = df[df[value_col].notna()].copy()
                if df.empty:
                    continue

                df = df.sort_values('日期', ascending=True).reset_index(drop=True)

                target_col_num = col_letter_to_num(target_col)

                for i, (_, row) in enumerate(df.iterrows()):
                    r = data_start_row + i
                    date_cell = ws.cell(row=r, column=date_col_num)
                    date_val = row['日期']
                    date_cell.value = date_val.to_pydatetime() if isinstance(date_val, pd.Timestamp) else date_val
                    date_cell.number_format = 'yyyy-mm-dd'

                    val = row[value_col]
                    if pd.notna(val):
                        cell = ws.cell(row=r, column=target_col_num)
                        cell.value = float(val)
                        cell.number_format = '0.00'

            print(f"    - 完成 {source_sheet} section [{date_col}] 共 {len(indicators)} 个指标")

    @staticmethod
    def media_internet_apply_yoy(context, params):
        """按 section 中指定的 yoy_col 写入月度同比公式 ={col}{row}/{col}{row-12}-1"""
        ws = context['ws']
        sheet_config = context['sheet_config']
        data_start_row = sheet_config.get('data_start_row', 11)

        sections = sheet_config.get('sections', [])
        for section in sections:
            yoy_col = section.get('yoy_col')
            if not yoy_col:
                continue

            indicators = section['indicators']
            if isinstance(indicators, dict):
                indicators = list(indicators.items())

            for _, value_col in indicators:
                value_col_num = col_letter_to_num(value_col)
                yoy_col_num = col_letter_to_num(yoy_col)

                for row in range(data_start_row, ws.max_row + 1):
                    source_cell = ws.cell(row=row, column=value_col_num)
                    if source_cell.value is None:
                        continue
                    prev_row = row - 12
                    if prev_row < data_start_row:
                        continue
                    prev_cell = ws.cell(row=prev_row, column=value_col_num)
                    if prev_cell.value is None:
                        continue
                    formula = f"={value_col}{row}/{value_col}{prev_row}-1"
                    cell = ws.cell(row=row, column=yoy_col_num)
                    cell.value = formula
                    cell.number_format = '0.00%'

            print(f"    - 完成同比公式写入: {[v for _, v in indicators]} -> {yoy_col}")
