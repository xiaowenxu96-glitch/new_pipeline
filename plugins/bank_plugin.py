import pandas as pd
import openpyxl
import re
from openpyxl.utils import column_index_from_string as column_letter_to_number, get_column_letter as column_number_to_letter
from datetime import datetime

# 为了重用之前的复杂代码，直接引入旧的 DataProcessor (作为示例平滑迁移，实际可全盘重构进这里)
from core_engine.data_processor import DataProcessor

class BankPlugin:
    @staticmethod
    def bank_write_data(context, params):
        print(f"开始处理银行数据")
        ws = context['ws']
        reader = context['data_reader']
        sheet_config = context['sheet_config']
        source_sheet = sheet_config['source_sheet']
        indicator_col_map = sheet_config['indicators']
        start_row = params.get('start_row', 50)
        start_column = params.get('start_column', 2)
        start_date = params.get('start_date', '2021-01-01')
        date_format = params.get('date_format', 'yyyy-mm')
        divide_by_100 = params.get('divide_by_100', False)
        start_date_ts = pd.to_datetime(start_date)

        # 1. 从缓存中获取所需所有指标数据
        indicator_dfs = {}
        for indicator in indicator_col_map:
            ind_data = reader.read_indicator_data(source_sheet, indicator)
            df = ind_data["data"].copy()

            if df.empty or '日期' not in df.columns:
                indicator_dfs[indicator] = pd.DataFrame(columns=['日期'])
                print(f"    - 工作表[{source_sheet}] 未找到指标 {indicator}，该列跳过写入")
                continue

            df['日期'] = pd.to_datetime(df['日期'], errors='coerce')

            # 只保留 start_date 及以后数据
            df = df[df['日期'].notna() & (df['日期'] >= start_date_ts)].copy()
            df = df.sort_values('日期', ascending=True).reset_index(drop=True)

            indicator_dfs[indicator] = df

        # 2. 使用第一个指标的数据作为基础日期表，并筛选季度数据（3、6、9、12月）
        first_indicator = list(indicator_col_map.keys())[0]
        first_df = indicator_dfs[first_indicator].copy()

        if first_df.empty:
            print(f"    - {source_sheet} 的第一个指标 {first_indicator} 无有效数据，跳过写入")
            return

        # 筛选只包含3、6、9、12月的数据
        quarterly_df = first_df[first_df['日期'].dt.month.isin([3, 6, 9, 12])].copy()
        
        if quarterly_df.empty:
            print(f"    - {source_sheet} 的第一个指标 {first_indicator} 中没有3、6、9、12月的数据，跳过写入")
            return

        base_df = quarterly_df.reset_index(drop=True)

        # 3. 写入统一日期列和指标数据（应用单位转换）
        for i, row in base_df.iterrows():
            current_row = start_row + i
            current_date = row['日期']

            if date_format == 'qQyy':
                # 提取季度和两位数年份，拼接成 1Q19 的字符串
                quarter = (current_date.month - 1) // 3 + 1
                year_str = current_date.strftime('%y')
                display_value = f"{quarter}Q{year_str}"
                
                date_cell = ws.cell(row=current_row, column=start_column, value=display_value)
                # 字符串不需要额外设置 number_format
            else:
                # 其他常规格式，按标准日期格式写入
                date_cell = ws.cell(row=current_row, column=start_column, value=current_date)
                date_cell.number_format = date_format
       
        # 4. 写入指标数据，全部按 base_df 的日期对齐（应用单位转换）
        base_date_list = base_df['日期'].tolist()

        for indicator, col in list(indicator_col_map.items()):
            df = indicator_dfs[indicator].copy()
            # 对每个指标也筛选季度数据
            df_quarterly = df[df['日期'].dt.month.isin([3, 6, 9, 12])].copy()
            
            if df_quarterly.empty:
                print(f"    - 指标 {indicator} 中没有3、6、9、12月的数据")
                continue
                
            value_col = df_quarterly.columns[-1]
            value_map = dict(zip(df_quarterly['日期'], df_quarterly[value_col]))
            col_num = column_letter_to_number(col) if isinstance(col, str) else int(col)    
            for i, date in enumerate(base_date_list):
                value = value_map.get(date, None)
                if value is not None and isinstance(value, (int, float)):
                    if divide_by_100:
                        value = value /100.0
                target_cell = ws.cell(
                    row=start_row + i,
                    column=col_num,
                    value=value
                )
                
                # 继承同列首行的原始格式
                original_cell = ws.cell(row=start_row, column=col_num)
                target_cell.number_format = original_cell.number_format

        print(f"    - 完成银行基础数据写入")


    @staticmethod
    def bank_commercial_write_data(context, params):
        ws = context['ws']
        reader = context['data_reader']
        sheet_config = context['sheet_config']
        source_sheet = sheet_config['source_sheet']
        indicator_col_map = sheet_config['indicators']
        start_row = params.get('start_row', 50)
        start_column = params.get('start_column', 1)
        unit_conversion_row = params.get('unit_conversion', 0)
        start_date = params.get('start_date', '2014-01-01')
        date_format = params.get('date_format', 'yyyy-mm')
        # 新增：支持多列日期写入，可以是单个列号或列号列表
        date_columns = params.get('date_columns', None)
        start_date_ts = pd.to_datetime(start_date)

        def convert_value(value, unit_conversion):
            """根据列和配置的倍率转换数值"""
            if value is None or not isinstance(value, (int, float)):
                return value
            if unit_conversion:
                return value * unit_conversion
            return value
                
        # 1. 从缓存中获取所需所有指标数据
        indicator_dfs = {}
        for indicator in indicator_col_map:
            ind_data = reader.read_indicator_data(source_sheet, indicator)
            df = ind_data["data"].copy()

            if df.empty or '日期' not in df.columns:
                indicator_dfs[indicator] = pd.DataFrame(columns=['日期'])
                print(f"    - 工作表[{source_sheet}] 未找到指标 {indicator}，该列跳过写入")
                continue

            df['日期'] = pd.to_datetime(df['日期'], errors='coerce')

            # 只保留 start_date 及以后数据
            df = df[df['日期'].notna() & (df['日期'] >= start_date_ts)].copy()
            df = df.sort_values('日期').reset_index(drop=True)

            indicator_dfs[indicator] = df

        # 2. 使用第一个指标的数据作为基础日期表（不再补全日期）
        first_indicator = list(indicator_col_map.keys())[0]
        first_df = indicator_dfs[first_indicator].copy()

        if first_df.empty:
            print(f"    - {source_sheet} 的第一个指标 {first_indicator} 无有效数据，跳过写入")
            return

        # 直接使用第一个指标的数据作为基础日期表
        base_df = first_df.copy()

        # 3. 写入统一日期列和指标数据（应用单位转换）
        if date_columns is not None:
            if isinstance(date_columns, int):
                date_cols_list = [date_columns]
            elif isinstance(date_columns, str):
                try:
                    date_cols_list = [column_letter_to_number(date_columns)]
                except:
                    date_cols_list = [int(date_columns)]
            else:
                date_cols_list = []
                for col in date_columns:
                    if isinstance(col, str):
                        try:
                            date_cols_list.append(column_letter_to_number(col))
                        except:
                            date_cols_list.append(int(col))
                    else:
                        date_cols_list.append(int(col))
        else:
            date_cols_list = [start_column]
        
        for i, row in base_df.iterrows():
            current_row = start_row + i
            current_date = row['日期']

            for col_num in date_cols_list:
                if col_num >= 1:
                    date_cell = ws.cell(row=current_row, column=col_num, value=current_date)
                    date_cell.number_format = date_format

        # 4. 写入指标数据，全部按 base_df 的日期对齐（应用单位转换）
        base_date_list = base_df['日期'].tolist()

        for indicator, col in list(indicator_col_map.items()):
            df = indicator_dfs[indicator].copy()
            value_col = df.columns[-1]
            value_map = dict(zip(df['日期'], df[value_col]))
            
            try:
                col_num = column_letter_to_number(col) if isinstance(col, str) else int(col)
                if col_num < 1:
                    print(f"    - 跳过列 {col}，计算出的列号 {col_num} 无效")
                    continue
                    
                if unit_conversion_row >= 1 and col_num >= 1:
                    unit_cell = ws.cell(row=unit_conversion_row, column=col_num)
                    if unit_cell.value is not None and unit_cell.value > 0:
                        unit_conversion = unit_cell.value
                    else:
                        unit_conversion = 1
                else:
                    unit_conversion = 1
            except Exception as e:
                print(f"    - 获取单位转换因子失败，使用默认值1: {e}")
                unit_conversion = 1
                
            for i, date in enumerate(base_date_list):
                value = value_map.get(date, None)
                converted_value = convert_value(value, unit_conversion)
                target_cell = ws.cell(
                    row=start_row + i,
                    column=col_num,
                    value=converted_value
                )
                
                original_cell = ws.cell(row=start_row, column=col_num)
                target_cell.number_format = original_cell.number_format
                
        return

    @staticmethod
    def bank_commercial_formula(context, params):
        ws = context['ws']

        if params.get('is_curve'):
            import math  # 引入 math 处理隐形的 NaN 空值
            start_row = params.get('start_row', 5)
            target_formula_template = params.get('target_formula', '={source_column}{row}/100')
            target_columns = params.get('target_column', [])
            source_columns = params.get('source_column', [])
            custom_format = params.get('format', '0.00%')
            max_row = ws.max_row
            
            for target_col, source_col in zip(target_columns, source_columns):
                target_col_num = column_letter_to_number(target_col) if isinstance(target_col, str) else int(target_col)
                for current_row in range(start_row, max_row + 1):
                    # 1. 检查 A 列时间列，跳过完全没有日期的空行
                    time_val = ws.cell(row=current_row, column=1).value
                    if time_val is None or str(time_val).strip() == "":
                        continue
                    
                    # 2. 究极源数据空值校验
                    source_val = ws[f"{source_col}{current_row}"].value
                    is_empty = False
                    
                    if source_val is None:
                        is_empty = True
                    elif isinstance(source_val, str) and str(source_val).strip() == "":
                        is_empty = True
                    elif isinstance(source_val, float) and math.isnan(source_val):
                        is_empty = True
                    # 如果需要把真实的数字 0 也当成空值跳过，可以解除下面这行的注释
                    # elif source_val == 0 or source_val == 0.0:
                    #     is_empty = True
                        
                    if is_empty:
                        # 如果源数据确实为空，强行给目标格留白，绝不写公式
                        target_cell = ws.cell(row=current_row, column=target_col_num)
                        target_cell.value = "" 
                        continue
                        
                    # 3. 只有确认有真实数据，才写入公式和格式
                    formula = target_formula_template.replace('{source_column}', source_col).replace('{row}', str(current_row))
                    target_cell = ws.cell(row=current_row, column=target_col_num)
                    target_cell.value = formula
                    if custom_format:
                        target_cell.number_format = custom_format
                        
            print("  -> [系统日志] 借壳执行利率曲线映射公式完毕（已部署究极空值过滤）。")
            return
        
        # 模式C：借壳执行连续求和（贷款余额SHEET
        source_starts = params.get('source_start', [])
        source_ends = params.get('source_end', [])
        if source_starts and source_ends:
            start_row = params.get('start_row', 3)
            target_formula_template = params.get('target_formula', '=SUM({source_start}{row}:{source_end}{row})')
            target_columns = params.get('target_column', [])
            custom_format = params.get('format', '#,##0.00')

            max_row = ws.max_row
            
            for target_col, src_start, src_end in zip(target_columns, source_starts, source_ends):
                target_col_num = column_letter_to_number(target_col) if isinstance(target_col, str) else int(target_col)
                
                # 改为 for 循环，绝不提前退出
                for current_row in range(start_row, max_row + 1):
                    time_val = ws.cell(row=current_row, column=2).value
                    # 遇到空行直接跳过，继续往下处理
                    if time_val is None or str(time_val).strip() == "":
                        continue
                        
                    formula = target_formula_template.format(row=current_row, source_start=src_start, source_end=src_end)
                    target_cell = ws.cell(row=current_row, column=target_col_num)
                    target_cell.value = formula
                    if custom_format:
                        target_cell.number_format = custom_format
                        
            print("  -> [系统日志] 借壳执行区间求和公式完毕（已屏蔽空行干扰）。")
            return
        
        # ==========================================
        # 模式 A：全新的动态加权模式 (通过识别 rules 触发)
        # ==========================================
        rules = params.get('rules', [])
        if rules:
            import math
            start_row = params.get('start_row', 3)
            ref_offset = params.get('ref_offset', 35)
            ref_sheet = params.get('ref_sheet', '贷款余额')
            custom_format = params.get('format', '0.00%')
            
            print(f"  -> [系统日志] 借壳执行动态加权公式，共 {len(rules)} 条规则...")
            
            for rule in rules:
                target_col = rule['target']
                source_cols = rule['sources']
                weight_cols = rule.get('weights', source_cols)
                
                target_col_num = column_letter_to_number(target_col) if isinstance(target_col, str) else int(target_col)
                current_row = start_row
                written_count = 0
                
                while True:
                    # 检查 B 列时间是否结束
                    time_val = ws.cell(row=current_row, column=2).value
                    if time_val is None or str(time_val).strip() == "":
                        break
                        
                    ref_row = current_row + ref_offset
                    active_sources = []
                    active_weights = []
                    
                    # 校验哪些列是真的有数据
                    for sc, wc in zip(source_cols, weight_cols):
                        cell_val = ws[f"{sc}{current_row}"].value
                        is_empty = False
                        
                        if cell_val is None:
                            is_empty = True
                        elif isinstance(cell_val, str) and str(cell_val).strip() == "":
                            is_empty = True
                        elif isinstance(cell_val, float) and math.isnan(cell_val):
                            is_empty = True
                            
                        if not is_empty:
                            active_sources.append(sc)
                            active_weights.append(wc)
                    
                    # 动态拼接并写入
                    if len(active_sources) == 0:
                        formula = ""
                    elif len(active_sources) == 1:
                        formula = f"={active_sources[0]}{current_row}"
                    else:
                        num_parts = [f"{sc}{current_row}*{ref_sheet}!{wc}{ref_row}" for sc, wc in zip(active_sources, active_weights)]
                        den_parts = [f"{ref_sheet}!{wc}{ref_row}" for wc in active_weights]
                        formula = f"=({'+'.join(num_parts)})/SUM({','.join(den_parts)})"
                        
                    target_cell = ws.cell(row=current_row, column=target_col_num)
                    target_cell.value = formula if formula else ""
                    if custom_format:
                        target_cell.number_format = custom_format
                        
                    current_row += 1
                    written_count += 1
                    
            print(f"  -> [系统日志] 动态加权全部写入完成。")
            return  # 动态模式执行完毕后，直接退出函数

        # ==========================================
        # 模式 B：兼容老配置的原有逻辑 (没有 rules 时触发)
        # ==========================================
        start_row = params.get('start_row', 6)
        target_formula_template = params.get('target_formula', '')
        target_columns = params.get('target_column', [])
        source_columns = params.get('source_column', [])
        ref_offset = params.get('ref_offset', 0)
        tname = params.get('tname', None)
        custom_format = params.get('format', None)
        
        if not target_formula_template or not target_columns:
            return

        if source_columns and len(target_columns) == len(source_columns):
            for target_col, source_col in zip(target_columns, source_columns):
                try:
                    target_col_num = column_letter_to_number(target_col) if isinstance(target_col, str) else int(target_col)
                    source_col_num = column_letter_to_number(source_col) if isinstance(source_col, str) else int(source_col)
                    
                    current_row = start_row
                    while True:
                        check_col_num = target_col_num - 1
                        check_cell = ws.cell(row=current_row, column=check_col_num if check_col_num >= 1 else target_col_num)
                        if check_cell.value is None:
                            break
                        
                        source_col_letter = column_number_to_letter(source_col_num)
                        formula = target_formula_template.replace('{source_column}', source_col_letter)
                        formula = re.sub(r'\{row-(\d+)\}', lambda m: str(current_row - int(m.group(1))), formula)
                        formula = formula.replace('{row}', str(current_row))
                        if tname:
                            formula = formula.replace('{tname}', tname)
                        
                        target_cell = ws.cell(row=current_row, column=target_col_num)
                        target_cell.value = formula
                        target_cell.number_format = custom_format or "General"
                        current_row += 1
                except Exception as e:
                    print(f"    - 处理列 {target_col} 时出错: {e}")
        else:
            # 模式 B 的另一半逻辑不变...
            for target_col in target_columns:
                try:
                    target_col_num = column_letter_to_number(target_col) if isinstance(target_col, str) else int(target_col)
                    current_row = start_row
                    while True:
                        check_col_num = target_col_num - 1
                        check_cell = ws.cell(row=current_row, column=check_col_num if check_col_num >= 1 else target_col_num)
                        if check_cell.value is None:
                            break
                        ref_row = current_row + ref_offset
                        prev_row = current_row - 1
                        formula = target_formula_template.format(row=current_row, ref_row=ref_row, prev_row=prev_row)
                        target_cell = ws.cell(row=current_row, column=target_col_num)
                        target_cell.value = formula
                        if custom_format:
                            target_cell.number_format = custom_format
                        current_row += 1
                except Exception as e:
                    pass
    
    @staticmethod
    def bank_bll_formula(context, params):
        """
        新增：处理多列区间连续求和公式 (适用于贷款余额等 Sheet 中的整体列汇总)
        支持 target_column, source_start, source_end 并行列表
        """
        ws = context['ws']
        start_row = params.get('start_row', 3)
        target_formula_template = params.get('target_formula', '=SUM({source_start}{row}:{source_end}{row})')
        target_columns = params.get('target_column', [])
        source_starts = params.get('source_start', [])
        source_ends = params.get('source_end', [])
        custom_format = params.get('format', '#,##0.00')

        if not target_columns or not source_starts or not source_ends:
            print("    - [警告] 缺少求和公式参数配置，跳过写入")
            return

        max_row = ws.max_row

        print(f"  ->[系统日志] 开始执行 bank_bll_formula 区间求和，共 {len(target_columns)} 列...")
        
        for target_col, src_start, src_end in zip(target_columns, source_starts, source_ends):
            try:
                target_col_num = column_letter_to_number(target_col) if isinstance(target_col, str) else int(target_col)
                written_count = 0

                # 强制循环到表格最底部
                for current_row in range(start_row, max_row + 1):
                    # 使用 B 列(时间列) 作为检测列
                    time_val = ws.cell(row=current_row, column=2).value
                    # 如果时间列为空，跳过该行，继续往下找（绝不使用 break 提前退出）
                    if time_val is None or str(time_val).strip() == "":
                        continue

                    # 渲染公式
                    formula = target_formula_template.format(
                        row=current_row,
                        source_start=src_start,
                        source_end=src_end
                    )

                    # 写入单元格
                    target_cell = ws.cell(row=current_row, column=target_col_num)
                    target_cell.value = formula
                    if custom_format:
                        target_cell.number_format = custom_format

                    written_count += 1

                print(f"    - [成功] 列 {target_col} 求和公式写入完毕（共 {written_count} 行）")
            except Exception as e:
                print(f"    - 处理列 {target_col} 求和公式出错: {e}")
        
    @staticmethod
    def bank_dynamic_weighted_formula(context, params):
        import math
        
        ws = context['ws']
        start_row = params.get('start_row', 3)
        ref_offset = params.get('ref_offset', 35)
        ref_sheet = params.get('ref_sheet', '贷款余额')
        custom_format = params.get('format', '0.00%')
        rules = params.get('rules', [])
        
        print(f"  -> [系统日志] 开始执行动态加权公式，共读取到 {len(rules)} 条列规则...")
        
        if not rules:
            print("    - [警告] 缺少 rules 配置，跳过写入")
            return
            
        for rule in rules:
            target_col = rule['target']
            source_cols = rule['sources']
            weight_cols = rule.get('weights', source_cols)
            
            target_col_num = column_letter_to_number(target_col) if isinstance(target_col, str) else int(target_col)
            current_row = start_row
            
            print(f"    - 正在计算目标列 {target_col} (源: {source_cols})")
            
            written_count = 0
            while True:
                # 检查 B列（时间列，column=2）是否为空来判断是否到底
                time_val = ws.cell(row=current_row, column=2).value
                if time_val is None or str(time_val).strip() == "":
                    print(f"      * 行号 {current_row} 的时间列(B列)为空，停止往下遍历。")
                    break
                    
                ref_row = current_row + ref_offset
                active_sources = []
                active_weights = []
                
                # 1. 扫描当前行真实的有效数据
                for sc, wc in zip(source_cols, weight_cols):
                    cell_val = ws[f"{sc}{current_row}"].value
                    
                    is_empty = False
                    if cell_val is None:
                        is_empty = True
                    elif isinstance(cell_val, str) and str(cell_val).strip() == "":
                        is_empty = True
                    elif isinstance(cell_val, float) and math.isnan(cell_val):
                        is_empty = True
                        
                    if not is_empty:
                        active_sources.append(sc)
                        active_weights.append(wc)
                
                # 2. 动态生成公式
                if len(active_sources) == 0:
                    formula = ""
                elif len(active_sources) == 1:
                    formula = f"={active_sources[0]}{current_row}"
                else:
                    num_parts = [f"{sc}{current_row}*{ref_sheet}!{wc}{ref_row}" for sc, wc in zip(active_sources, active_weights)]
                    den_parts = [f"{ref_sheet}!{wc}{ref_row}" for wc in active_weights]
                    formula = f"=({'+'.join(num_parts)})/SUM({','.join(den_parts)})"
                    
                # 3. 写入单元格
                target_cell = ws.cell(row=current_row, column=target_col_num)
                target_cell.value = formula if formula else ""
                
                if custom_format:
                    target_cell.number_format = custom_format
                    
                current_row += 1
                written_count += 1
                
            print(f"    - [成功] 列 {target_col} 写入完成，共填入 {written_count} 行公式。")

    @staticmethod
    def bank_curve_formula(context, params):
        ws = context['ws']
        start_row = params.get('start_row', 5)
        target_formula_template = params.get('target_formula', '={source_column}{row}/100')
        target_columns = params.get('target_column', [])
        source_columns = params.get('source_column', [])
        custom_format = params.get('format', '0.00%')

        if not target_columns or not source_columns:
            print("    - [警告] 缺少 target_column 或 source_column 配置")
            return

        max_row = ws.max_row
        
        # 将 target (如 R) 和 source (如 D) 一一对应遍历
        for target_col, source_col in zip(target_columns, source_columns):
            target_col_num = column_letter_to_number(target_col) if isinstance(target_col, str) else int(target_col)
            written_count = 0
            
            # 强制遍历到表格底部，跳过空行
            for current_row in range(start_row, max_row + 1):
                # 检查 B 列时间列，为空则跳过 (防空行中断)
                time_val = ws.cell(row=current_row, column=2).value
                if time_val is None or str(time_val).strip() == "":
                    continue
                    
                # 渲染公式：替换为如 =D5/100
                formula = target_formula_template.replace('{source_column}', source_col).replace('{row}', str(current_row))
                
                target_cell = ws.cell(row=current_row, column=target_col_num)
                target_cell.value = formula
                if custom_format:
                    target_cell.number_format = custom_format
                
                written_count += 1
                
        print("  -> [系统日志] 利率曲线映射公式写入完毕。")
   
ACTION_REGISTRY = {
    "bank_write_data": BankPlugin.bank_write_data,
    "bank_commercial_write_data": BankPlugin.bank_commercial_write_data,
    "bank_commercial_formula": BankPlugin.bank_commercial_formula,
    "bank_bll_formula": BankPlugin.bank_bll_formula,
    "bank_dynamic_weighted_formula": BankPlugin.bank_dynamic_weighted_formula,
    "bank_curve_formula": BankPlugin.bank_curve_formula,
}