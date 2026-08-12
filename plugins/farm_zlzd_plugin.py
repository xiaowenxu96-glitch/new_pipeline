import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string as column_letter_to_number, get_column_letter as column_number_to_letter
from datetime import datetime

from core_engine.data_processor import DataProcessor

class FarmPlugin:

    @staticmethod
    def farm_write_data(context, params):
        """
        功能说明：
            处理农业基础指标数据，写入到目标工作表中。
            与 farm_export_plugin 的区别：支持 date_col 参数指定日期写入列。
            从 data_reader 缓存中读取指标数据，以第一个指标的日期为基准，
            将所有指标数据按此日期对齐写入 Excel。
            日期按升序排列（最早日期在前），使用日度数据格式。

        params:
            context: dict — Pipeline 上下文，包含 ws、data_reader、sheet_config。
            params: dict — 从 YAML action 配置中解析的参数，包含：
                - start_row: int (默认 11) — 数据写入的起始行号。
                - start_date: str (默认 '2014-01-01') — 数据最早日期。
                - date_format: str (默认 'yyyy-mm-dd') — 日期列的数字格式。
                - date_col: str (默认 'A') — 日期写入的目标列。
                - indicators: dict — 指标编码到列号的映射（从 action 配置中读取）。
        """
        ws = context['ws']
        reader = context['data_reader']
        sheet_config = context['sheet_config']
        source_sheet = sheet_config['source_sheet']
        indicator_col_map = params.get('indicators') or {}
        # 防御：YAML 中冒号后缺少空格时可能被解析为字符串而非字典
        if isinstance(indicator_col_map, str):
            print(f"    - 错误：indicators 被解析为字符串 '{indicator_col_map}'，"
                  f"请检查 YAML 配置中冒号后是否有空格")
            return
        if not isinstance(indicator_col_map, dict) or not indicator_col_map:
            print(f"    - 错误：indicators 无效（类型: {type(indicator_col_map).__name__}），跳过写入")
            return

        start_row = params.get('start_row', 11)
        start_date = params.get('start_date', '2014-01-01')
        date_format = params.get('date_format', 'yyyy-mm-dd')
        start_date_ts = pd.to_datetime(start_date)

        # 解析 date_col：日期写入的目标列，默认 A 列
        date_col_raw = params.get('date_col', 'A')
        try:
            date_col_num = column_letter_to_number(date_col_raw) if isinstance(date_col_raw, str) else int(date_col_raw)
        except Exception:
            date_col_num = 1

        # 1. 从缓存中获取所需所有指标数据
        indicator_dfs = {}
        for indicator in indicator_col_map:
            ind_data = reader.read_indicator_data(source_sheet, indicator)
            df = ind_data["data"].copy()

            if df.empty or '日期' not in df.columns:
                indicator_dfs[indicator] = pd.DataFrame(columns=['日期'])
                print(f"    - 工作表[{source_sheet}] 未找到指标 {indicator}，该列跳过写入")
                continue

            df['日期'] = pd.to_datetime(df['日期'], errors='coerce')

            # 只保留 start_date 及以后数据
            df = df[df['日期'].notna() & (df['日期'] >= start_date_ts)].copy()
            df = df.sort_values('日期', ascending=True).reset_index(drop=True)

            indicator_dfs[indicator] = df

        # 2. 使用第一个指标的数据作为基础日期表
        first_indicator = list(indicator_col_map.keys())[0]
        first_df = indicator_dfs[first_indicator].copy()

        if first_df.empty:
            print(f"    - {source_sheet} 的第一个指标 {first_indicator} 无有效数据，跳过写入")
            return

        base_df = first_df.copy()
        last_data_row = start_row + len(base_df) - 1

        # 0. 取消写入区域内已有的合并单元格，避免 MergedCell 只读报错
        col_nums = set()
        for col in indicator_col_map.values():
            try:
                col_nums.add(column_letter_to_number(col) if isinstance(col, str) else int(col))
            except Exception:
                pass
        all_cols = {date_col_num} | col_nums  # 加上日期列
        for merged_range in list(ws.merged_cells.ranges):
            if (merged_range.min_row >= start_row or merged_range.max_row >= start_row) and \
               any(merged_range.min_col <= c <= merged_range.max_col for c in all_cols):
                ws.unmerge_cells(str(merged_range))

        # 3. 写入统一日期列
        for i, row in base_df.iterrows():
            current_row = start_row + i
            current_date = row['日期']

            date_cell = ws.cell(row=current_row, column=date_col_num, value=current_date)
            date_cell.number_format = date_format

        # 4. 写入指标数据，全部按 base_df 的日期对齐
        base_date_list = base_df['日期'].tolist()

        for indicator, col in list(indicator_col_map.items()):
            df = indicator_dfs[indicator].copy()
            value_col = df.columns[-1]
            value_map = dict(zip(df['日期'], df[value_col]))

            try:
                col_num = column_letter_to_number(col) if isinstance(col, str) else int(col)
                if col_num < 1:
                    print(f"    - 跳过列 {col}，计算出的列号 {col_num} 无效")
                    continue
            except Exception as e:
                print(f"    - 解析列号失败，跳过列 {col}: {e}")
                continue

            for i, date in enumerate(base_date_list):
                value = value_map.get(date, None)
                ws.cell(
                    row=start_row + i,
                    column=col_num,
                    value=value
                )

        print(f"    - 完成农业基础数据写入")

        # 5. 处理公式（如果配置了 formulas）
        formulas = params.get('formulas', [])
        if formulas:
            for fm_cfg in formulas:
                col_letter = fm_cfg.get('col', '')
                template = fm_cfg.get('template', '')
                fm_start_row = fm_cfg.get('start_row', start_row + 1)
                if not col_letter or not template:
                    continue
                try:
                    col_num = column_letter_to_number(col_letter)
                except Exception:
                    print(f"    - 公式列号解析失败: {col_letter}，跳过")
                    continue
                for r in range(fm_start_row, last_data_row + 1):
                    expr = template.replace('{row}', str(r)).replace('{prev_row}', str(r - 1))
                    ws.cell(row=r, column=col_num, value=expr)
                print(f"    - 公式写入完成: 列 {col_letter}，行 {fm_start_row}~{last_data_row}")

        # 6. 清除下方残留旧数据，避免图表坐标轴出现 1900 日期
        clear_start = last_data_row + 1
        clear_end = 2000  # 足够覆盖图表引用范围

        # 清除日期列残留
        for r in range(clear_start, clear_end + 1):
            ws.cell(row=r, column=date_col_num, value=None)

        # 清除各指标列残留
        for col in indicator_col_map.values():
            try:
                col_num = column_letter_to_number(col) if isinstance(col, str) else int(col)
                for r in range(clear_start, clear_end + 1):
                    ws.cell(row=r, column=col_num, value=None)
            except Exception:
                pass

        print(f"    - 已清除第 {clear_start}-{clear_end} 行残留数据")

    @staticmethod
    def farm_compute_period_avg(context, params):
        """
        功能说明：
            在已有日度数据行旁边追加周均值和月均值列。
            周均值写入该周最后一天所在行，月均值写入该月最后一天所在行。

        params:
            context: dict — Pipeline 上下文，包含 ws。
            params: dict — 从 YAML action 配置中解析的参数，包含：
                - date_col: str (默认 'A') — 日期所在的列。
                - data_start_row: int (默认 5) — 数据起始行号。
                - week_avg_cols: dict — {源列: 目标列}，计算指定源列的周均值并写入目标列。
                - month_avg_cols: dict — {源列: 目标列}，计算指定源列的月均值并写入目标列。
        """
        ws = context['ws']

        date_col_raw = params.get('date_col', 'A')
        try:
            date_col_num = column_letter_to_number(date_col_raw) if isinstance(date_col_raw, str) else int(date_col_raw)
        except Exception:
            date_col_num = 1

        data_start_row = params.get('data_start_row', 5)

        week_avg_map = params.get('week_avg_cols') or {}
        month_avg_map = params.get('month_avg_cols') or {}

        if not week_avg_map and not month_avg_map:
            print("    - 未配置 week_avg_cols 或 month_avg_cols，跳过周期均值计算")
            return

        for name, m in [('week_avg_cols', week_avg_map), ('month_avg_cols', month_avg_map)]:
            if isinstance(m, str):
                print(f"    - 错误：{name} 被解析为字符串 '{m}'，请检查 YAML 中冒号后是否有空格")
                return

        # 1. 扫描数据行，收集日期和各源列数值
        records = []
        row = data_start_row
        all_src_cols = set()
        for m in [week_avg_map, month_avg_map]:
            all_src_cols.update(m.keys())

        while True:
            date_val = ws.cell(row=row, column=date_col_num).value
            if date_val is None:
                break
            if isinstance(date_val, datetime):
                rec = {'row': row, 'date': date_val}
                for src_col in all_src_cols:
                    src_col_num = column_letter_to_number(src_col)
                    rec[src_col] = ws.cell(row=row, column=src_col_num).value
                records.append(rec)
            row += 1

        if not records:
            print("    - 未找到数据行，跳过周期均值计算")
            return

        df = pd.DataFrame(records)
        last_data_row = records[-1]['row']

        # 收集所有目标列，用于后续清除残留
        all_tgt_col_nums = set()
        for m in [week_avg_map, month_avg_map]:
            for tgt_col in m.values():
                try:
                    all_tgt_col_nums.add(column_letter_to_number(tgt_col))
                except Exception:
                    pass

        # 取消目标列区域的合并单元格
        for merged_range in list(ws.merged_cells.ranges):
            if merged_range.min_row >= data_start_row and \
               any(merged_range.min_col <= c <= merged_range.max_col for c in all_tgt_col_nums):
                ws.unmerge_cells(str(merged_range))

        # 2. 计算周均值（按 ISO 周年分组）
        if week_avg_map:
            df['iso_year'] = df['date'].apply(lambda d: d.isocalendar()[0])
            df['iso_week'] = df['date'].apply(lambda d: d.isocalendar()[1])

            for src_col, tgt_col in week_avg_map.items():
                tgt_col_num = column_letter_to_number(tgt_col)
                grouped = df.groupby(['iso_year', 'iso_week'])
                for (_year, _week), group in grouped:
                    valid_vals = group[src_col].dropna()
                    if len(valid_vals) == 0:
                        continue
                    avg_val = valid_vals.mean()
                    last_row = int(group.loc[group['date'].idxmax(), 'row'])
                    ws.cell(row=last_row, column=tgt_col_num, value=avg_val)

            print(f"    - 周均值计算完成，写入列: {list(week_avg_map.values())}")

        # 3. 计算月均值
        if month_avg_map:
            df['year'] = df['date'].dt.year
            df['month'] = df['date'].dt.month

            for src_col, tgt_col in month_avg_map.items():
                tgt_col_num = column_letter_to_number(tgt_col)
                grouped = df.groupby(['year', 'month'])
                for (_year, _month), group in grouped:
                    valid_vals = group[src_col].dropna()
                    if len(valid_vals) == 0:
                        continue
                    avg_val = valid_vals.mean()
                    last_row = int(group.loc[group['date'].idxmax(), 'row'])
                    ws.cell(row=last_row, column=tgt_col_num, value=avg_val)

            print(f"    - 月均值计算完成，写入列: {list(month_avg_map.values())}")

        # 4. 清除目标列下方残留旧数据
        clear_start = last_data_row + 1
        clear_end = 2000
        for tgt_col_num in all_tgt_col_nums:
            for r in range(clear_start, clear_end + 1):
                ws.cell(row=r, column=tgt_col_num, value=None)

        print(f"    - 已清除目标列第 {clear_start}-{clear_end} 行残留数据")
