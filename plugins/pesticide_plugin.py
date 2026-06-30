import pandas as pd
from datetime import datetime
from collections import defaultdict
from openpyxl.utils import column_index_from_string as col_letter_to_num


class PesticidePlugin:
    """农药 Pipeline 插件：按 AA 指标代码读取源数据并写入目标文档"""

    @staticmethod
    def pesticide_write_sheet(context, params):
        ws = context['ws']
        reader = context['data_reader']
        sheet_config = context['sheet_config']
        defaults = context['defaults']

        source_sheet = sheet_config['source_sheet']
        data_start_row = sheet_config.get('data_start_row') or defaults.get('start_row', 2)
        date_col_letter = sheet_config.get('date_col', 'A')
        date_col_num = col_letter_to_num(date_col_letter)
        indicators = sheet_config['indicators']

        # 收集所有指标的日期→值映射
        indicator_maps = {}  # target_col_num → {date: value}
        all_dates = set()

        for aa_code, target_col in indicators.items():
            ind_data = reader.read_indicator_data(source_sheet, aa_code)
            df = ind_data['data']
            if df.empty:
                continue

            df = df[df['日期'].notna()].copy()
            value_col = df.columns[-1]
            df = df[df[value_col].notna()].copy()
            if df.empty:
                continue

            date_val_map = {}
            for _, row in df.iterrows():
                d = row['日期']
                v = row[value_col]
                if pd.notna(d) and pd.notna(v):
                    date_val_map[d] = float(v)

            tgt_col_num = col_letter_to_num(target_col)
            indicator_maps[tgt_col_num] = date_val_map
            all_dates.update(date_val_map.keys())

        if not indicator_maps:
            print(f"    - [{source_sheet}] 未找到任何指标数据")
            return

        # 按日期升序排列（最早→最新）
        date_list = sorted(all_dates)

        # 写日期列
        for i, date_val in enumerate(date_list):
            row = data_start_row + i
            c = ws.cell(row=row, column=date_col_num)
            c.value = date_val.to_pydatetime() if isinstance(date_val, pd.Timestamp) else date_val
            c.number_format = 'yyyy-mm-dd'

        # 写各指标值
        for tgt_col_num, date_map in indicator_maps.items():
            for i, date_val in enumerate(date_list):
                val = date_map.get(date_val)
                if val is not None:
                    c = ws.cell(row=data_start_row + i, column=tgt_col_num)
                    c.value = val
                    c.number_format = '0.00'

        print(f"    - [{source_sheet}] 写入 {len(date_list)} 行, {len(indicator_maps)} 个指标")

    # ================================================================
    # 汇总 Sheet 填充：季度均价、年度均价、年初价格
    # ================================================================
    @staticmethod
    def pesticide_write_summary(context, params):
        wb = context['wb']
        ws = context['ws']
        sc = context['sheet_config']

        data_sheet_name = sc['source_data_sheet']
        data_start_row = sc.get('data_start_row', 4)

        if data_sheet_name not in wb.sheetnames:
            print(f"    !! 数据 sheet [{data_sheet_name}] 不存在")
            return
        data_ws = wb[data_sheet_name]

        # --- 1. 从数据 sheet 读取所有品种的日期→值映射 ---
        product_data = {}  # name → {date: value}
        product_cols = {}
        all_dates = set()

        for c in range(2, data_ws.max_column + 1):
            pname = data_ws.cell(row=2, column=c).value
            if not pname:
                continue
            pname = str(pname).strip()
            product_cols[c] = pname
            date_vals = {}
            for r in range(data_start_row, data_ws.max_row + 1):
                dv = data_ws.cell(row=r, column=1).value
                v = data_ws.cell(row=r, column=c).value
                if dv is None or v is None:
                    continue
                if isinstance(dv, datetime):
                    try:
                        date_vals[dv] = float(v)
                    except (ValueError, TypeError):
                        pass
            if date_vals:
                product_data[pname] = date_vals
                all_dates.update(date_vals.keys())

        if not product_data:
            print(f"    - [{data_sheet_name}] 汇总: 无数据")
            return

        all_dates = sorted(all_dates)
        latest_date = all_dates[-1]

        # --- 2. 确定最近完整季度和完整年份 ---
        # 最近完整季度 = 最新数据所在季度的上一个季度
        latest_q = (latest_date.month - 1) // 3 + 1
        if latest_q == 1:
            last_complete_q = 4
            last_complete_q_year = latest_date.year - 1
        else:
            last_complete_q = latest_q - 1
            last_complete_q_year = latest_date.year

        # 最近完整年份 = 最新数据年份 - 1（当前年份未完成）
        last_complete_year = latest_date.year - 1

        # --- 3. 生成 12 个季度标签 ---
        quarters = []
        y, q = last_complete_q_year, last_complete_q
        for _ in range(12):
            quarters.append((y, q))
            q -= 1
            if q == 0:
                q = 4
                y -= 1
        quarters.reverse()  # 从最早到最新

        # --- 4. 生成 7 个年度标签 ---
        years = list(range(last_complete_year - 6, last_complete_year + 1))

        # --- 5. 确定最近完整月份 ---
        if latest_date.month == 1:
            last_complete_month = 12
            last_complete_month_year = latest_date.year - 1
        else:
            last_complete_month = latest_date.month - 1
            last_complete_month_year = latest_date.year

        # --- 6. 在汇总 sheet Row 2 定位各区块列号 ---
        quarterly_start = None
        annual_start = None
        year_price_col = None
        month_end_col = None
        quarter_end_col = None

        for c in range(1, ws.max_column + 1):
            r1 = ws.cell(row=1, column=c).value
            r2 = ws.cell(row=2, column=c).value
            r1s = str(r1).strip() if r1 else ''
            r2s = str(r2).strip() if r2 else ''
            if '季度均价' in r1s and quarterly_start is None:
                quarterly_start = c
            if '年度均价' in r1s and annual_start is None:
                annual_start = c
            if ('初价格' in r2s or '年末价格' in r2s) and year_price_col is None:
                year_price_col = c
            if '月末价格' in r2s and month_end_col is None:
                month_end_col = c
            if ('季末价格' in r2s or '季初价格' in r2s) and quarter_end_col is None:
                quarter_end_col = c

        # --- 7. 更新 Row 2 标签 ---
        # 季度标签（12 列）
        if quarterly_start:
            for i in range(12):
                ws.cell(row=2, column=quarterly_start + i).value = f'{quarters[i][0]}Q{quarters[i][1]}'

        # 年度标签（7 列）
        if annual_start:
            for i, yr in enumerate(years):
                ws.cell(row=2, column=annual_start + i).value = str(yr)
        # 年初价格标签
        target_ys_year = last_complete_year
        if year_price_col:
            ws.cell(row=2, column=year_price_col).value = f'{target_ys_year}年初价格'
        # 上个月月末价格标签
        if month_end_col:
            ws.cell(row=2, column=month_end_col).value = f'{last_complete_month}月末价格'
        # 上个季度季末价格标签
        if quarter_end_col:
            ws.cell(row=2, column=quarter_end_col).value = f'{last_complete_q_year}Q{last_complete_q}季末价格'

        # --- 8. 逐品种计算并写入 ---
        quarterly_written = 0
        annual_written = 0
        ys_written = 0
        month_end_written = 0
        quarter_end_written = 0

        for r in range(3, ws.max_row + 1):
            product = ws.cell(row=r, column=1).value
            if not product:
                continue
            product = str(product).strip()
            if product not in product_data:
                continue

            pdata = product_data[product]

            # 季度均价
            if quarterly_start:
                for i, (y, q) in enumerate(quarters):
                    q_vals = [v for d, v in pdata.items()
                              if d.year == y and (d.month - 1) // 3 + 1 == q]
                    if q_vals:
                        avg = sum(q_vals) / len(q_vals)
                        cell = ws.cell(row=r, column=quarterly_start + i)
                        cell.value = round(avg, 4)
                        cell.number_format = '0.00'
                        quarterly_written += 1

            # 年度均价
            if annual_start:
                for i, yr in enumerate(years):
                    y_vals = [v for d, v in pdata.items() if d.year == yr]
                    if y_vals:
                        avg = sum(y_vals) / len(y_vals)
                        cell = ws.cell(row=r, column=annual_start + i)
                        cell.value = round(avg, 4)
                        cell.number_format = '0.00'
                        annual_written += 1

            # 年初价格
            if year_price_col:
                ys_vals = [(d, v) for d, v in pdata.items() if d.year == target_ys_year]
                if ys_vals:
                    ys_vals.sort(key=lambda x: x[0])
                    cell = ws.cell(row=r, column=year_price_col)
                    cell.value = ys_vals[0][1]
                    cell.number_format = '0.00'
                    ys_written += 1

            # 上个月月末价格（最后完整月的最后一条数据）
            if month_end_col:
                me_vals = [(d, v) for d, v in pdata.items()
                           if d.year == last_complete_month_year and d.month == last_complete_month]
                if me_vals:
                    me_vals.sort(key=lambda x: x[0])
                    cell = ws.cell(row=r, column=month_end_col)
                    cell.value = me_vals[-1][1]
                    cell.number_format = '0.00'
                    month_end_written += 1

            # 上个季度季末价格（最后完整季度的最后一条数据）
            if quarter_end_col:
                qe_vals = [(d, v) for d, v in pdata.items()
                           if d.year == last_complete_q_year and (d.month - 1) // 3 + 1 == last_complete_q]
                if qe_vals:
                    qe_vals.sort(key=lambda x: x[0])
                    cell = ws.cell(row=r, column=quarter_end_col)
                    cell.value = qe_vals[-1][1]
                    cell.number_format = '0.00'
                    quarter_end_written += 1

        # --- 9. 清除溢出列的旧标签和旧数据 ---
        if quarterly_start:
            for c in range(quarterly_start + 12, min(quarterly_start + 20, ws.max_column + 1)):
                r2v = ws.cell(row=2, column=c).value
                if r2v and 'Q' in str(r2v):
                    for r in range(3, ws.max_row + 1):
                        ws.cell(row=r, column=c).value = None
                    ws.cell(row=2, column=c).value = None
        if annual_start:
            for c in range(annual_start + 7, min(annual_start + 15, ws.max_column + 1)):
                r2v = ws.cell(row=2, column=c).value
                if r2v:
                    try:
                        if 2000 <= int(str(r2v).strip()) <= 2050:
                            for r in range(3, ws.max_row + 1):
                                ws.cell(row=r, column=c).value = None
                            ws.cell(row=2, column=c).value = None
                    except ValueError:
                        pass

        print(f"    - [{data_sheet_name}→{sc['sheet_name']}] "
              f"季度:{quarterly_written} 年度:{annual_written} "
              f"年初:{ys_written} 月末:{month_end_written} 季末:{quarter_end_written}")
