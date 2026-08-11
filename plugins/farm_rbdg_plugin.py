import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string as column_letter_to_number, get_column_letter as column_number_to_letter
from datetime import datetime

from core_engine.data_processor import DataProcessor

class FarmPlugin:
    
    @staticmethod
    def farm_write_data(context, params):
        """
        功能说明：
            处理农业基础指标数据，写入到目标工作表中。
            从 data_reader 缓存中读取指标数据，以第一个指标的日期为基准，
            将所有指标数据按此日期对齐写入 Excel。
            日期按升序排列（最早日期在前），使用日度数据格式。

        params:
            context: dict — Pipeline 上下文，包含 ws、data_reader、sheet_config。
            params: dict — 从 YAML action 配置中解析的参数，包含：
                - start_row: int (默认 11) — 数据写入的起始行号。
                - start_date: str (默认 '2014-01-01') — 数据最早日期。
                - date_format: str (默认 'yyyy-mm-dd') — 日期列的数字格式。
                - indicators: dict — 指标编码到列号的映射（从 action 配置中读取）。
        """
        ws = context['ws']
        reader = context['data_reader']
        sheet_config = context['sheet_config']
        source_sheet = sheet_config['source_sheet']
        indicator_col_map = params.get('indicators') or {}
        # 防御：YAML 中冒号后缺少空格时可能被解析为字符串而非字典
        if isinstance(indicator_col_map, str):
            print(f"    - 错误：indicators 被解析为字符串 '{indicator_col_map}'，"
                  f"请检查 YAML 配置中冒号后是否有空格")
            return
        if not isinstance(indicator_col_map, dict) or not indicator_col_map:
            print(f"    - 错误：indicators 无效（类型: {type(indicator_col_map).__name__}），跳过写入")
            return

        start_row = params.get('start_row', 11)
        start_date = params.get('start_date', '2014-01-01')
        date_format = params.get('date_format', 'yyyy-mm-dd')
        start_date_ts = pd.to_datetime(start_date)

        # 1. 从缓存中获取所需所有指标数据
        indicator_dfs = {}
        for indicator in indicator_col_map:
            ind_data = reader.read_indicator_data(source_sheet, indicator)
            df = ind_data["data"].copy()

            if df.empty or '日期' not in df.columns:
                indicator_dfs[indicator] = pd.DataFrame(columns=['日期'])
                print(f"    - 工作表[{source_sheet}] 未找到指标 {indicator}，该列跳过写入")
                continue

            df['日期'] = pd.to_datetime(df['日期'], errors='coerce')

            # 只保留 start_date 及以后数据
            df = df[df['日期'].notna() & (df['日期'] >= start_date_ts)].copy()
            df = df.sort_values('日期', ascending=True).reset_index(drop=True)

            indicator_dfs[indicator] = df

        # 2. 使用第一个指标的数据作为基础日期表
        first_indicator = list(indicator_col_map.keys())[0]
        first_df = indicator_dfs[first_indicator].copy()

        if first_df.empty:
            print(f"    - {source_sheet} 的第一个指标 {first_indicator} 无有效数据，跳过写入")
            return

        base_df = first_df.copy()

        # 3. 写入统一日期列
        for i, row in base_df.iterrows():
            current_row = start_row + i
            current_date = row['日期']

            date_cell = ws.cell(row=current_row, column=1, value=current_date)
            date_cell.number_format = date_format

        # 4. 写入指标数据，全部按 base_df 的日期对齐
        base_date_list = base_df['日期'].tolist()

        for indicator, col in list(indicator_col_map.items()):
            df = indicator_dfs[indicator].copy()
            value_col = df.columns[-1]
            value_map = dict(zip(df['日期'], df[value_col]))

            try:
                col_num = column_letter_to_number(col) if isinstance(col, str) else int(col)
                if col_num < 1:
                    print(f"    - 跳过列 {col}，计算出的列号 {col_num} 无效")
                    continue
            except Exception as e:
                print(f"    - 解析列号失败，跳过列 {col}: {e}")
                continue

            for i, date in enumerate(base_date_list):
                value = value_map.get(date, None)
                ws.cell(
                    row=start_row + i,
                    column=col_num,
                    value=value
                )

        print(f"    - 完成农业基础数据写入")

        # 5. 处理公式（如果配置了 formulas）
        last_data_row = start_row + len(base_date_list) - 1
        formulas = params.get('formulas', [])
        if formulas:
            for fm_cfg in formulas:
                col_letter = fm_cfg.get('col', '')
                template = fm_cfg.get('template', '')
                fm_start_row = fm_cfg.get('start_row', start_row + 1)
                if not col_letter or not template:
                    continue
                try:
                    col_num = column_letter_to_number(col_letter)
                except Exception:
                    print(f"    - 公式列号解析失败: {col_letter}，跳过")
                    continue
                for r in range(fm_start_row, last_data_row + 1):
                    expr = template.replace('{row}', str(r)).replace('{prev_row}', str(r - 1))
                    ws.cell(row=r, column=col_num, value=expr)
                print(f"    - 公式写入完成: 列 {col_letter}，行 {fm_start_row}~{last_data_row}")

        # 6. 清除下方残留旧数据，避免图表坐标轴出现 1900 日期
        clear_start = last_data_row + 1
        clear_end = 2000  # 足够覆盖图表引用范围

        # 清除 A 列（日期）残留
        for r in range(clear_start, clear_end + 1):
            ws.cell(row=r, column=1, value=None)

        # 清除各指标列残留
        for col in indicator_col_map.values():
            try:
                col_num = column_letter_to_number(col) if isinstance(col, str) else int(col)
                for r in range(clear_start, clear_end + 1):
                    ws.cell(row=r, column=col_num, value=None)
            except Exception:
                pass

        print(f"    - 已清除第 {clear_start}-{clear_end} 行残留数据")
    
    @staticmethod
    def farm_write_linechart(context, params):
        """
        功能说明：
            根据工作表已有数据创建折线图。
            数据第一列为时间（X 轴），第 2-n 列为数据系列。
            数据不包含表头，折线图不需要图例或标签。
            图表外框无线条，背景无 Y 轴横线参考线。

        params:
            context: dict — Pipeline 上下文，包含 ws、wb、sheet_config。
            params: dict — 从 YAML action 配置中解析的参数，包含：
                - data_range: str (必需) — 数据范围，例如 "A5:B670"。
                第一列为时间轴，其余列为数据系列。
                - y_axis_min: float (可选) — Y 轴最小值，例如 5。
                - y_axis_max: float (可选) — Y 轴最大值，例如 12。
                - year_range: int (可选) — 仅展示最近 N 年的数据。
                读取第一个数据行（A 列）的日期，向前推算 N 年作为 X 轴起点，
                自动裁剪数据范围，只保留该时间段内的数据。
                例如 year_range=5，A5 日期为 2026/07/02，则 X 轴范围为 2021/07/02 ~ 2026/07/02。
                - chart_position: str (可选) — 图表左上角锚点，例如 "D5"。
                默认放在数据右侧第 2 列、与数据同行。
                - chart_width: int (可选，默认 20) — 图表宽度（单位近似列宽）。
                - chart_height: int (可选，默认 12) — 图表高度（单位近似行高）。
        """
        import re
        from datetime import datetime
        from dateutil.relativedelta import relativedelta
        from openpyxl.chart import LineChart, Reference

        ws = context['ws']

        # 1. 解析数据范围
        data_range = params.get('data_range', '')
        if not data_range:
            print("    - 未指定 data_range，跳过图表创建")
            return

        m = re.match(
            r'^\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)$', data_range
        )
        if not m:
            print(f"    - 无法解析 data_range: {data_range}，跳过图表创建")
            return

        col_start_str, row_start_str, col_end_str, row_end_str = m.groups()
        row_start = int(row_start_str)
        row_end = int(row_end_str)
        col_start_num = column_letter_to_number(col_start_str)
        col_end_num = column_letter_to_number(col_end_str)

        # year_range processing - 确定数据行的起止范围
        # 数据按时间升序排列（最早日期在 row_start，最新在 row_end）
        actual_row_start = row_start  # 实际使用的起始行
        year_range = params.get('year_range')

        if year_range and year_range > 0:
            # 获取最后一行（最新日期）的日期值
            last_date_cell = ws.cell(row=row_end, column=col_start_num).value

            if last_date_cell:
                # 处理日期值（可能是 datetime 或 date 类型）
                if isinstance(last_date_cell, datetime):
                    latest_date = last_date_cell
                else:
                    try:
                        latest_date = datetime.strptime(str(last_date_cell), '%Y-%m-%d')
                    except:
                        latest_date = last_date_cell

                # 计算 X 轴起点日期（最新日期向前推 N 年）
                if isinstance(latest_date, datetime):
                    start_date = latest_date - relativedelta(years=year_range)

                    # 在数据行中从前往后查找第一个 >= start_date 的行
                    # 数据按时间升序排列，找到第一个不早于 start_date 的行作为起始行
                    for row_idx in range(row_start, row_end + 1):
                        cell_value = ws.cell(row=row_idx, column=col_start_num).value

                        if cell_value:
                            # 转换为 datetime 进行比较
                            if isinstance(cell_value, datetime):
                                current_date = cell_value
                            else:
                                try:
                                    current_date = datetime.strptime(str(cell_value), '%Y-%m-%d')
                                except:
                                    continue

                            # 找到第一个 >= start_date 的数据行（进入范围了）
                            if current_date >= start_date:
                                actual_row_start = row_idx
                                break

                    print(f"    - year_range={year_range}: 最新日期={latest_date.strftime('%Y-%m-%d')}, "
                        f"起始日期={start_date.strftime('%Y-%m-%d')}, 数据行范围: 行{actual_row_start}~行{row_end}")
                else:
                    print(f"    - 无法解析最后一行日期: {last_date_cell}，使用全部数据范围")
            else:
                print(f"    - 最后一行日期为空，使用全部数据范围")

        # 2. 创建折线图
        chart = LineChart()
        chart.style = 2

        # X 轴类别引用（第一列 = 时间）
        cats = Reference(
            ws,
            min_col=col_start_num,
            max_col=col_start_num,
            min_row=actual_row_start,
            max_row=row_end,
        )

        # 数据系列（第 2-n 列），每列作为一条折线
        for col_num in range(col_start_num + 1, col_end_num + 1):
            data_ref = Reference(
                ws,
                min_col=col_num,
                max_col=col_num,
                min_row=actual_row_start,
                max_row=row_end,
            )
            chart.add_data(data_ref, from_rows=False, titles_from_data=False)

        # 为所有系列统一设置类别（X 轴）
        chart.set_categories(cats)

        # 3. Y 轴取值范围
        y_axis_min = params.get('y_axis_min')
        y_axis_max = params.get('y_axis_max')
        if y_axis_min is not None:
            chart.y_axis.scaling.min = y_axis_min
        if y_axis_max is not None:
            chart.y_axis.scaling.max = y_axis_max

        # 4. 样式：无图例，无标题，无线条背景
        chart.legend = None
        chart.title = None

        # 去掉系列名称
        for series in chart.series:
            series.tx = None

        # 移除 Y 轴水平网格线（背景横线参考线）
        chart.y_axis.majorGridlines = None
        chart.y_axis.minorGridlines = None

        # 移除图表外框线条（无线条）
        _set_chart_no_border(chart)

        # 5. 图表位置
        chart_pos = params.get('chart_position', '')
        if chart_pos:
            pos_m = re.match(r'^\$?([A-Z]+)\$?(\d+)$', chart_pos)
            if pos_m:
                chart.anchor = chart_pos
            else:
                default_col = column_number_to_letter(col_end_num + 2)
                chart.anchor = f'{default_col}{row_start}'
        else:
            default_col = column_number_to_letter(col_end_num + 2)
            chart.anchor = f'{default_col}{row_start}'

        chart.width = params.get('chart_width', 15)
        chart.height = params.get('chart_height', 8)

        ws.add_chart(chart)
        
        if actual_row_start != row_start:
            print(f"    - 折线图创建完成，数据范围: 行{actual_row_start}~行{row_end}（原始范围: {data_range}），共 {col_end_num - col_start_num} 条折线")
        else:
            print(f"    - 折线图创建完成，数据范围: {data_range}，共 {col_end_num - col_start_num} 条折线")

        return


def _set_chart_no_border(chart):
    """
    移除图表外框线条（无线条）和背景填充。
    通过设置 chartSpace 的 graphicalProperties 实现：
    - 外框线条 → 无 (<a:ln><a:noFill/></a:ln>)
    - 背景填充 → 无 (<a:noFill/>)
    """
    try:
        from openpyxl.chart.shapes import GraphicalProperties
        from openpyxl.drawing.line import LineProperties

        gp = GraphicalProperties()
        gp.noFill = True  # 背景无填充

        line_props = LineProperties()
        line_props.noFill = True  # 外框无线条
        gp.ln = line_props

        chart.graphicalProperties = gp
    except Exception:
        pass  # 若 openpyxl 版本不兼容则静默跳过，不影响图表创建