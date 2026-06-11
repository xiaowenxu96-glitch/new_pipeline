import re
import zipfile
import shutil
import pandas as pd
from openpyxl.utils import column_index_from_string as col_letter_to_num


class BaijiuPlugin:
    """白酒 Pipeline 插件：按产品名匹配写入 + 图表 XML 更新"""

    # ================================================================
    # 数据写入：源列→目标列（按产品名匹配目标表头）
    # ================================================================
    @staticmethod
    def baijiu_write_sheet(context, params):
        ws = context['ws']
        reader = context['data_reader']
        sc = context['sheet_config']

        source_sheet = sc['source_sheet']
        data_start = sc.get('data_start_row', 9)
        date_col_num = col_letter_to_num(sc['date_col'])

        df = reader.read_sheet_data(source_sheet)

        # --- 1. 读取源产品名 → source_col_idx ---
        meta = df.iloc[:10]
        src_products = {}  # product_name → 0-based df col index
        src_brands = {}
        for c in range(1, len(df.columns)):
            full = str(meta.iloc[0, c]) if pd.notna(meta.iloc[0, c]) else ''
            parts = full.split(':')
            brand = parts[3] if len(parts) > 3 else ''
            pname = parts[4] if len(parts) > 4 else ''
            if pname:
                src_products[pname] = c
                src_brands[pname] = brand

        # --- 2. 读取目标表头 → target_col_num ---
        tgt_headers = {}  # header_text → target_col_num
        for c in range(date_col_num + 1, ws.max_column + 1):
            hdr = ws.cell(row=sc.get('header_read_row', data_start - 2), column=c).value
            if hdr:
                tgt_headers[str(hdr).strip()] = c
            # Also check row above (brand row)
            hdr2 = ws.cell(row=sc.get('header_read_row', data_start - 2) - 1, column=c).value
            if hdr2:
                tgt_headers[str(hdr2).strip()] = c

        # --- 3. 手动映射表 ---
        # Config 格式：{源产品名: 目标表头}
        raw_map = sc.get('manual_map', {})
        manual_col_map = {}  # tgt_col_num → src_col_idx
        for src_pname, tgt_hdr in raw_map.items():
            # 查找源列
            src_idx = src_products.get(src_pname)
            if src_idx is None:
                continue
            # 查找目标列
            tgt_col = tgt_headers.get(tgt_hdr)
            if tgt_col is None:
                continue
            manual_col_map[tgt_col] = src_idx

        # --- 4. 匹配：目标表头 → 源列索引 ---
        col_map = dict(manual_col_map)  # 先手动映射
        used_src = set(col_map.values())

        # 跳过日期/品牌行
        data_headers = [(k, v) for k, v in tgt_headers.items()
                        if not any(kw in k for kw in ['日期', '从早', '从新'])]

        # 第一遍：精确名称匹配
        rest = []
        for tgt_name, tgt_col in data_headers:
            if tgt_col in col_map:
                continue
            if tgt_name in src_products and src_products[tgt_name] not in used_src:
                col_map[tgt_col] = src_products[tgt_name]
                used_src.add(src_products[tgt_name])
            else:
                rest.append((tgt_name, tgt_col))

        # 第二遍：模糊匹配
        for tgt_name, tgt_col in rest:
            if tgt_col in col_map:
                continue
            for sp, sc_idx in src_products.items():
                if sc_idx in used_src:
                    continue
                if sp in tgt_name or tgt_name in sp:
                    col_map[tgt_col] = sc_idx
                    used_src.add(sc_idx)
                    break

        print(f"    - [{source_sheet}] {len(col_map)}列匹配")

        # --- 5. 数据处理 ---
        data_df = df.iloc[10:].reset_index(drop=True)
        if data_df.empty:
            return

        data_df.columns = [f'col_{i}' for i in range(len(data_df.columns))]
        data_df = data_df.rename(columns={'col_0': 'date'})
        data_df['date'] = pd.to_datetime(data_df['date'], errors='coerce')
        data_df = data_df[data_df['date'].notna()].copy()
        data_df = data_df.sort_values('date', ascending=True).reset_index(drop=True)

        # --- 6. 写日期 ---
        for i, dv in enumerate(data_df['date']):
            c = ws.cell(row=data_start + i, column=date_col_num)
            c.value = dv.to_pydatetime() if isinstance(dv, pd.Timestamp) else dv
            c.number_format = 'yyyy-mm-dd'

        # --- 7. 写数据 ---
        for tgt_col, src_col_idx in col_map.items():
            src_col_name = f'col_{src_col_idx}'
            if src_col_name not in data_df.columns:
                continue
            for i, val in enumerate(data_df[src_col_name]):
                if pd.notna(val):
                    c = ws.cell(row=data_start + i, column=tgt_col)
                    try:
                        c.value = float(val)
                    except (ValueError, TypeError):
                        c.value = val
                    c.number_format = '0.00'

    # ================================================================
    # 图表 XML 更新
    # ================================================================
    _CHART_SHEET_MAP = {
        10: 0, 11: 0, 12: 0,
        13: 1,
        15: 2, 16: 2, 17: 2, 18: 2, 19: 2, 20: 2, 21: 2, 22: 2, 23: 2,
        24: 2, 25: 2, 26: 2, 27: 2, 28: 2, 29: 2, 30: 2, 31: 2,
        32: 3, 33: 3, 34: 3, 35: 3, 36: 3, 37: 3, 38: 3, 39: 3, 40: 3,
        41: 3, 42: 3, 43: 3, 44: 3, 45: 3, 46: 3, 47: 3, 48: 3,
    }

    @staticmethod
    def baijiu_finalize_charts(context, params):
        filepath = context['filepath']
        config = context['config']
        tmp = filepath + '.tmp'

        with zipfile.ZipFile(filepath, 'r') as zin, zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename.startswith('xl/charts/chart') and item.filename.endswith('.xml'):
                    data = BaijiuPlugin._patch_chart_xml(data, item.filename, filepath, config)
                zout.writestr(item, data)

        shutil.move(tmp, filepath)
        print(f"    [XML] 图表更新完成")

    @staticmethod
    def _patch_chart_xml(xml_bytes, filename, filepath, config):
        xml = xml_bytes.decode('utf-8')
        m = re.search(r'chart(\d+)\.xml', filename)
        if not m:
            return xml_bytes
        cn = int(m.group(1))

        sheet_idx = BaijiuPlugin._CHART_SHEET_MAP.get(cn)
        if sheet_idx is None:
            return xml_bytes
        sc = config['sheets'][sheet_idx]
        is_first_two = (sheet_idx in [0, 1])

        cs, er = BaijiuPlugin._chart_rows(filepath, sc, is_first_two)
        if cs is None:
            return xml_bytes

        if sc.get('max_data_cols', 0):
            max_tc = col_letter_to_num(sc['date_col']) + sc['max_data_cols']
            xml = BaijiuPlugin._strip_oob(xml, max_tc)

        xml = re.sub(
            r'\x24([A-Z]+)\x24(\d+):\x24([A-Z]+)\x24(\d+)',
            lambda m: f'${m.group(1)}${cs}:${m.group(3)}${er}',
            xml
        )

        if not is_first_two:
            xml = BaijiuPlugin._set_axis(xml)

        for tag in ['dLbls', 'numCache', 'strCache']:
            xml = re.sub(f'<(?:c:)?{tag}>.*?</(?:c:)?{tag}>', '', xml, flags=re.DOTALL)
            xml = re.sub(f'<(?:c:)?{tag}[^/>]*/>', '', xml)

        if cn == 10:
            xml = re.sub(r'<(?:c:)?min val="[^"]*"/>', '', xml)
            xml = re.sub(r'<(?:c:)?max val="[^"]*"/>', '', xml)

        return xml.encode('utf-8')

    @staticmethod
    def _strip_oob(xml, max_tc):
        pat = r'\x24([A-Z]+)\x24(\d+):\x24([A-Z]+)\x24(\d+)'
        def replacer(m):
            all_m = re.findall(pat, m.group(0))
            if all_m:
                try:
                    if col_letter_to_num(all_m[-1][0]) > max_tc:
                        return ''
                except:
                    pass
            return m.group(0)
        return re.sub(r'<ser>.*?</ser>', replacer, xml, flags=re.DOTALL)

    @staticmethod
    def _set_axis(xml):
        if '<majorUnit' in xml:
            xml = re.sub(r'<majorUnit val="[^"]*"/>', '<majorUnit val="365"/>', xml)
        else:
            xml = re.sub(r'<dateAx>', r'<dateAx><majorUnit val="365"/><minorUnit val="30"/>', xml)
        if '<minorUnit' not in xml:
            xml = re.sub(r'(<majorUnit[^/]+/>)', r'\1<minorUnit val="30"/>', xml)
        return xml

    @staticmethod
    def _chart_rows(filepath, sc, is_first_two):
        data_start = sc.get('data_start_row', 9)
        dc = col_letter_to_num(sc['date_col'])

        df = pd.read_excel(filepath, sheet_name=sc['sheet_name'], header=None)
        end_row, latest = data_start, None
        for r in range(len(df), data_start - 1, -1):
            v = df.iloc[r - 1, dc - 1]
            if pd.notna(v):
                if end_row == data_start:
                    end_row = r
                if latest is None:
                    try:
                        latest = pd.to_datetime(v)
                    except:
                        pass
                if end_row != data_start and latest is not None:
                    break

        if latest is None:
            return None, None

        if is_first_two:
            cutoff = pd.Timestamp(year=latest.year - 2, month=1, day=1)
            chart_start = data_start
            for r in range(data_start, end_row + 1):
                v = df.iloc[r - 1, dc - 1]
                if pd.notna(v):
                    try:
                        if pd.to_datetime(v) >= cutoff:
                            chart_start = r
                            break
                    except:
                        pass
        else:
            chart_start = data_start

        return chart_start, end_row
