import pandas as pd
from datetime import datetime, timedelta
from collections import OrderedDict
from openpyxl.utils import column_index_from_string as col_letter_to_num, get_column_letter as num_to_col_letter


class PetrochemicalPlugin:
    """石油化工 Pipeline 插件：写数 + 周均价/周同比 + 月均价/月环比 + 年均价/年同比"""

    @staticmethod
    def _build_date_val_map(reader, source_sheet, aa_code):
        ind_data = reader.read_indicator_data(source_sheet, aa_code)
        df = ind_data['data']
        if df.empty:
            return {}
        df = df[df['日期'].notna()].copy()
        value_col = df.columns[-1]
        df = df[df[value_col].notna()].copy()
        if df.empty:
            return {}
        out = {}
        for _, row in df.iterrows():
            d = row['日期']
            v = row[value_col]
            if pd.notna(d) and pd.notna(v):
                dt = d.to_pydatetime() if isinstance(d, pd.Timestamp) else d
                out[dt.replace(hour=0, minute=0, second=0, microsecond=0)] = float(v)
        return out

    @staticmethod
    def _excel_serial(dt):
        """日期转 Excel 序列号"""
        return (dt - datetime(1899, 12, 30)).days

    # ---- 写数 ----

    @staticmethod
    def petrochemical_write_sheet(context, params):
        """
        写入日度数据，保持自然降序（最新在前）。
        """
        ws = context['ws']
        reader = context['data_reader']
        sheet_config = context['sheet_config']
        defaults = context.get('defaults', {})

        source_sheet = sheet_config['source_sheet']
        data_start_row = (
            sheet_config.get('data_start_row')
            or sheet_config.get('start_row')
            or defaults.get('start_row', 2)
        )

        sections = sheet_config.get('sections', [])
        if not sections:
            sections = [{
                'date_col': sheet_config.get('date_col', 'A'),
                'data_start_row': data_start_row,
                'indicators': sheet_config.get('indicators', {}),
            }]

        date_to_row_map = {}

        for section in sections:
            indicators = section.get('indicators', {})
            if not indicators:
                continue

            sec_start_row = section.get('data_start_row', data_start_row)
            sec_date_col = section.get('date_col', 'A')
            date_col_num = col_letter_to_num(sec_date_col)

            for aa_code, target_col in indicators.items():
                date_val_map = PetrochemicalPlugin._build_date_val_map(reader, source_sheet, aa_code)
                if not date_val_map:
                    continue

                tgt_col_num = col_letter_to_num(target_col)
                all_dates = sorted(date_val_map.keys(), reverse=True)

                for i, date_val in enumerate(all_dates):
                    row = sec_start_row + i
                    c = ws.cell(row=row, column=date_col_num)
                    c.value = date_val
                    c.number_format = 'yyyy-mm-dd'

                    val = date_val_map.get(date_val)
                    if val is not None:
                        c = ws.cell(row=row, column=tgt_col_num)
                        c.value = val
                        c.number_format = '0.00'

                    date_to_row_map[date_val] = row

        context['_petro_date_to_row'] = date_to_row_map
        context['_petro_data_start_row'] = data_start_row
        print(f"    - [{source_sheet}] 写入 {len(date_to_row_map)} 行数据")

    # ---- 周均价 / 周环比 ----

    @staticmethod
    def petrochemical_apply_weekly_formulas(context, params):
        """
        以「上周日 + 本周一二三四」为一周，按锚定周四分组。
        区间从该周的 Thursday(最小行号) 到 Sunday(最大行号)，
        保证覆盖完整 5 天（节假日缺失的天恰好不在区间内）。
        """
        ws = context['ws']
        date_to_row_map = context.get('_petro_date_to_row', {})
        price_col = params.get('price_col', 'B')
        avg_col = col_letter_to_num(params.get('avg_col', 'C'))
        wow_col = col_letter_to_num(params.get('wow_col', 'D'))
        avg_col_letter = num_to_col_letter(avg_col)

        if not date_to_row_map:
            print("    - [周公式] 无日期映射，跳过")
            return

        # 自定义周：上周日 + 本周一二三四 = 一周
        # 周键 = 锚定周四日期
        week_groups = OrderedDict()
        for d, row in date_to_row_map.items():
            wd = d.weekday()
            if wd == 6:    # 周日 -> 下周周四
                anchor_thu = d + timedelta(days=4)
            elif wd <= 3:  # 周一~周四 -> 本周周四
                anchor_thu = d + timedelta(days=3 - wd)
            else:          # 周五~周六 -> 下周四
                anchor_thu = d + timedelta(days=3 - wd + 7)
            key = anchor_thu.strftime('%Y-%m-%d')
            if key not in week_groups:
                week_groups[key] = []
            week_groups[key].append(row)

        # 按周降序
        sorted_weeks = sorted(week_groups.items(), key=lambda x: x[0], reverse=True)
        week_records = []  # [(thu_row, sun_row)]
        for key, rows in sorted_weeks:
            rows_sorted = sorted(rows)
            thu_row = rows_sorted[0]   # 最小行号 = Thursday (降序Thu在上)
            sun_row = rows_sorted[-1]  # 最大行号 = Sunday (降序Sun在下)
            week_records.append((thu_row, sun_row))

        # 写 C/D 公式
        # 降序：idx=0 最新，idx 越大越旧
        # D 列：本周/上周-1，所以分母是下一周(idx+1)
        for idx, (thu_row, sun_row) in enumerate(week_records):
            # C 列：周均价 = AVERAGE(Thu行:Sun行)
            ws.cell(row=thu_row, column=avg_col).value = f"=AVERAGE({price_col}{thu_row}:{price_col}{sun_row})"
            ws.cell(row=thu_row, column=avg_col).number_format = '0.00'

            if idx + 1 < len(week_records):
                next_thu, _ = week_records[idx + 1]  # 下一周 = 更旧 = 行号更大
                # D 列第1行(周四行)：日价格周同比 = B{本周}/B{下周}-1
                ws.cell(row=thu_row, column=wow_col).value = f"={price_col}{thu_row}/{price_col}{next_thu}-1"
                ws.cell(row=thu_row, column=wow_col).number_format = '0.00%'

                # D 列第2行(周三行)：周均价环比 = C{本周}/C{下周}-1
                if thu_row + 1 <= ws.max_row:
                    ws.cell(row=thu_row + 1, column=wow_col).value = f"={avg_col_letter}{thu_row}/{avg_col_letter}{next_thu}-1"
                    ws.cell(row=thu_row + 1, column=wow_col).number_format = '0.00%'

        print(f"    - [周公式] 写入 {len(week_records)} 周均价/环比")

    # ---- 月均价 / 月环比 ----

    @staticmethod
    def petrochemical_apply_monthly_formulas(context, params):
        """
        按 (年,月) 分组（数据降序，同月行连续），找到每月 min/max 行：
        - F 列：Excel serial date（该月第一天）
        - G 列：=AVERAGE(B{min}:B{max})
        - H 列：=G{row}/G{prev_row}-1
        写入到每月第一天对应行。
        """
        ws = context['ws']
        date_to_row_map = context.get('_petro_date_to_row', {})
        price_col = params.get('price_col', 'B')
        month_date_col = col_letter_to_num(params.get('month_col', 'F'))
        avg_col = col_letter_to_num(params.get('month_avg_col', 'G'))
        mom_col = col_letter_to_num(params.get('month_mom_col', 'H'))
        avg_col_letter = num_to_col_letter(avg_col)

        if not date_to_row_map:
            print("    - [月公式] 无日期映射，跳过")
            return

        month_groups = OrderedDict()
        for d, row in date_to_row_map.items():
            key = (d.year, d.month)
            if key not in month_groups:
                month_groups[key] = {'rows': [], 'first_date': None, 'first_row': None}
            month_groups[key]['rows'].append(row)
            if month_groups[key]['first_date'] is None or d > month_groups[key]['first_date']:
                month_groups[key]['first_date'] = d
                month_groups[key]['first_row'] = row

        sorted_months = sorted(month_groups.items(), key=lambda x: (x[0][0], x[0][1]), reverse=True)
        month_records = [(info['first_date'], sorted(info['rows'])[0], sorted(info['rows'])[-1])
                         for key, info in sorted_months]

        # 集中写入：从 row 2 开始连续排列
        base_row = 2
        written = 0
        for idx, (first_date, start_r, end_r) in enumerate(month_records):
            row = base_row + idx

            # F：月日期
            month_first = datetime(first_date.year, first_date.month, 1)
            ws.cell(row=row, column=month_date_col).value = PetrochemicalPlugin._excel_serial(month_first)
            ws.cell(row=row, column=month_date_col).number_format = 'yyyy-mm-dd'

            # G：月均价
            ws.cell(row=row, column=avg_col).value = f"=AVERAGE({price_col}{start_r}:{price_col}{end_r})"
            ws.cell(row=row, column=avg_col).number_format = '0.00'

            # H：月环比 = G{本月}/G{下月}-1
            if idx + 1 < len(month_records):
                next_row = base_row + idx + 1
                ws.cell(row=row, column=mom_col).value = f"={avg_col_letter}{row}/{avg_col_letter}{next_row}-1"
                ws.cell(row=row, column=mom_col).number_format = '0.00%'

            written += 1

        print(f"    - [月公式] 写入 {written} 月均价/环比 (rows {base_row}-{base_row+written-1})")

    # ---- 年均价 / 年同比 ----

    @staticmethod
    def petrochemical_apply_yearly_formulas(context, params):
        """
        按年份分组（数据降序，同年行连续），找到每年 min/max 行：
        - J 列：年份
        - K 列：=AVERAGE(B{min}:B{max})
        - L 列：=K{row}/K{prev_row}-1
        写入到每年第一天对应行。
        """
        ws = context['ws']
        date_to_row_map = context.get('_petro_date_to_row', {})
        price_col = params.get('price_col', 'B')
        year_col = col_letter_to_num(params.get('year_col', 'J'))
        avg_col = col_letter_to_num(params.get('year_avg_col', 'K'))
        yoy_col = col_letter_to_num(params.get('year_yoy_col', 'L'))
        avg_col_letter = num_to_col_letter(avg_col)

        if not date_to_row_map:
            print("    - [年公式] 无日期映射，跳过")
            return

        year_groups = OrderedDict()
        for d, row in date_to_row_map.items():
            yr = d.year
            if yr not in year_groups:
                year_groups[yr] = {'rows': [], 'first_date': None, 'first_row': None}
            year_groups[yr]['rows'].append(row)
            if year_groups[yr]['first_date'] is None or d > year_groups[yr]['first_date']:
                year_groups[yr]['first_date'] = d
                year_groups[yr]['first_row'] = row

        sorted_years = sorted(year_groups.items(), key=lambda x: x[0], reverse=True)
        year_records = [(info['first_date'], sorted(info['rows'])[0], sorted(info['rows'])[-1])
                        for yr, info in sorted_years]

        # 集中写入：从 row 2 开始连续排列
        base_row = 2
        written = 0
        for idx, (first_date, start_r, end_r) in enumerate(year_records):
            row = base_row + idx

            # J：年份
            ws.cell(row=row, column=year_col).value = first_date.year

            # K：年均价
            ws.cell(row=row, column=avg_col).value = f"=AVERAGE({price_col}{start_r}:{price_col}{end_r})"
            ws.cell(row=row, column=avg_col).number_format = '0.00'

            # L：年同比 = K{本年}/K{下年}-1
            if idx + 1 < len(year_records):
                next_row = base_row + idx + 1
                ws.cell(row=row, column=yoy_col).value = f"={avg_col_letter}{row}/{avg_col_letter}{next_row}-1"
                ws.cell(row=row, column=yoy_col).number_format = '0.00%'

            written += 1

        print(f"    - [年公式] 写入 {written} 年均价/同比 (rows {base_row}-{base_row+written-1})")
