import pandas as pd
import re
import datetime
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.datetime import from_excel as datetime_from_excel

class PoPlugin:
    """
    建筑装饰行业数据插件，支持横向（宽表）数据写入:

    源数据格式（纵向）:
        日期       | 指标值
    2009-09-29    | 6.36
    2009-10-29    | 2.56

    目标模板写入后格式（横向）:
        A列: 指标名称 | B列起: 2009.9 | 2009.10 | ...
        第3行: 中国建筑月度数据 | 6.36   | 2.56   | ...
    """

    @staticmethod
    def PO_write_data(context, params):
        """
        横向写入PO数据：
        1. 从所有指标的源数据中提取统一的时间周期序列
        2. 将时间周期作为列头写入目标模板
        3. 将每个指标的数据按日期对号写入对应列

        参数:
            date_header_row: 日期列头写入行号 (默认1)
            data_start_column: 数据起始列号 (默认2)
            start_date: 数据起始日期，此前的数据将被过滤
        """
        ws = context['ws']
        reader = context['data_reader']
        sheet_config = context['sheet_config']
        source_sheet = sheet_config['source_sheet']
        indicator_row_map = sheet_config['indicators']  # code -> target_row

        date_header_row = params.get('date_header_row', 1)
        data_start_column = params.get('data_start_column', 2)
        start_date = params.get('start_date', '2020-01-01')
        start_date_ts = pd.to_datetime(start_date)

        # ---------------------------------------------------------------
        # 1. 读取所有指标数据，提取统一的时间周期序列
        # ---------------------------------------------------------------
        indicator_dfs = {}
        all_periods = set()

        for indicator_code in indicator_row_map:
            ind_data = reader.read_indicator_data(source_sheet, indicator_code)
            df = ind_data["data"].copy()

            if df.empty or '日期' not in df.columns:
                print(f"    - 无数据，跳过: {indicator_code}")
                continue

            df['日期'] = pd.to_datetime(df['日期'], errors='coerce')
            df = df[df['日期'].notna() & (df['日期'] >= start_date_ts)].copy()

            if df.empty:
                print(f"    - 无有效数据: {indicator_code}")
                continue

            indicator_dfs[indicator_code] = df

            for _, row in df.iterrows():
                all_periods.add((row['日期'].year, row['日期'].month))

        if not all_periods:
            print(f"    - 所有指标均无有效数据，跳过写入")
            return

        sorted_periods = sorted(all_periods)  # [(2009,9), (2009,10), ...]
        print(f"    - 共 {len(sorted_periods)} 个时间周期, 从 {sorted_periods[0][0]}.{sorted_periods[0][1]} 到 {sorted_periods[-1][0]}.{sorted_periods[-1][1]}")

        # ---------------------------------------------------------------
        # 2. 写入日期列头
        # ---------------------------------------------------------------
        PoPlugin._unmerge_row(ws, date_header_row)
        for i, (year, month) in enumerate(sorted_periods):
            col = data_start_column + i
            ws.cell(row=date_header_row, column=col, value=f"{year}.{month}")

        # ---------------------------------------------------------------
        # 3. 逐一写入指标数据
        # ---------------------------------------------------------------
        for indicator_code, target_row in indicator_row_map.items():
            if indicator_code not in indicator_dfs:
                continue

            df = indicator_dfs[indicator_code]
            value_col = df.columns[-1]

            # 构建 (year,month) -> value 映射
            value_map = {}
            for _, row in df.iterrows():
                value_map[(row['日期'].year, row['日期'].month)] = row[value_col]

            # 解除目标行合并单元格
            PoPlugin._unmerge_row(ws, target_row)

            write_count = 0
            for i, (year, month) in enumerate(sorted_periods):
                key = (year, month)
                if key in value_map:
                    ws.cell(row=target_row, column=data_start_column + i, value=value_map[key])
                    write_count += 1

            print(f"    - 指标 {indicator_code} -> 第{target_row}行, 写入 {write_count} 个数据点")

        print(f"    - 完成PO数据写入")

    @staticmethod
    def _unmerge_row(ws, target_row):
        """解除覆盖目标行的所有合并单元格，以便写入数据"""
        for merged_range in list(ws.merged_cells.ranges):
            if merged_range.min_row <= target_row <= merged_range.max_row:
                ws.unmerge_cells(str(merged_range))

    @staticmethod
    def _parse_period(period_str: str):
        """将 '2020.1' / '2020.01' 解析为 (year, month)"""
        match = re.match(r'(\d{4})[.](\d{1,2})', period_str)
        if match:
            return int(match.group(1)), int(match.group(2))
        return None

    @staticmethod
    def PO_write_formula(context, params):
        """
        Write formulas horizontally from start_column to the last non-empty column in date_row.

        Supported placeholders:
        {col}         -> current column letter (A, B, C...)
        {col-N}       -> column letter N positions to the left
        {row}         -> current row number
        {source_row}  -> source data row number (from params, e.g. 4, 7, 10...)
        """
        ws = context['ws']
        start_row = params.get('start_row')
        start_column = params.get('start_column', 2)
        date_row = params.get('date_row')
        formula_template = params.get('target_formula', '')
        custom_format = params.get('format', None)
        source_row = params.get('source_row', None)

        if not formula_template or not start_row or not date_row:
            print("    - missing required params, skip formula write")
            return

        # Determine end column by scanning date_row
        col = start_column
        end_column = start_column - 1
        while True:
            if ws.cell(row=date_row, column=col).value is None:
                end_column = col - 1
                break
            col += 1

        if end_column < start_column:
            print("    - no data in date_row, skip formula write")
            return

        # Find the max offset in {col-N} to skip invalid columns
        offset_pat = re.compile(r'\{col-(\d+)\}')
        max_offset = 0
        for m in offset_pat.finditer(formula_template):
            max_offset = max(max_offset, int(m.group(1)))

        write_count = 0
        for current_col in range(start_column, end_column + 1):
            if max_offset > 0 and current_col - max_offset < start_column:
                continue

            formula = formula_template

            formula = offset_pat.sub(
                lambda m: get_column_letter(current_col - int(m.group(1))),
                formula
            )
            formula = formula.replace('{col}', get_column_letter(current_col))
            formula = formula.replace('{row}', str(start_row))
            if source_row is not None:
                formula = formula.replace('{source_row}', str(source_row))

            cell = ws.cell(row=start_row, column=current_col)
            if isinstance(cell, MergedCell):
                # skip cells that are part of a merged range
                continue
            cell.value = formula
            if custom_format:
                cell.number_format = custom_format
            write_count += 1

        end_letter = get_column_letter(end_column)
        print(f"    - wrote {write_count} formulas, row {start_row}, cols {get_column_letter(max(start_column, start_column+max_offset))}-{end_letter}")