import re
import pandas as pd
from openpyxl.utils import column_index_from_string


class DecorationPlugin:

    @staticmethod
    def _render_template(template, row):
        """支持 {row}, {prev_row}, {row+N}, {row-N} 等偏移语法。"""
        def replacer(m):
            key = m.group(1)
            if key == 'prev_row':
                return str(row - 1)
            parts = re.split(r'([+-])', key)
            if len(parts) == 3 and parts[0] == 'row':
                op = parts[1]
                offset = int(parts[2])
                return str(row + offset if op == '+' else row - offset)
            return m.group(0)
        return re.sub(r'\{([^}]+)\}', replacer, template)

    @staticmethod
    def decoration_write_sheet(context, params):
        """
    将建筑装饰的指标数据从源工作表写入目标工作表。

    通过 DataReader 读取 sheet_config['indicators'] 中配置的指标，
    按日期对齐后写入目标列，并清除下方残留数据。

    上下文：
        ws            — openpyxl 目标工作表对象
        data_reader   — 已缓存源文件的 DataReader 实例
        sheet_config  — 字典：{sheet_name, source_sheet, indicators, actions}
    参数：
        start_row     — 数据起始行（默认 2）
        date_col      — 日期所在列字母（默认 'A'）
        start_date    — 数据最早日期（默认 '2014-01-01'）
        date_format   — 日期单元格的 Excel 数字格式（默认 'yyyy-mm-dd'）
        """
        ws = context['ws']
        reader = context['data_reader']
        sheet_config = context['sheet_config']
        source_sheet = sheet_config['source_sheet']
        defaults = context.get('defaults', {})

        # ── 1. Extract indicator mapping from sheet config ──
        indicators_raw = sheet_config.get('indicators', {})
        indicator_col_map = {}
        if isinstance(indicators_raw, dict):
            for k, v in indicators_raw.items():
                if isinstance(v, dict):
                    indicator_col_map.update(v)
                elif isinstance(v, str):
                    indicator_col_map[k] = v

        if not indicator_col_map:
            print("    - 未找到指标配置，跳过写入")
            return

        start_row = params.get('start_row', defaults.get('start_row', 2))
        date_col = params.get('date_col', 'A')
        start_date = params.get('start_date', defaults.get('start_date', '2014-01-01'))
        date_format = params.get('date_format', defaults.get('date_format', 'yyyy-mm-dd'))
        start_date_ts = pd.to_datetime(start_date)
        date_col_num = column_index_from_string(date_col)

        formulas = params.get('formulas', [])

        if not indicator_col_map and not formulas:
            print("    - 未找到指标配置和公式配置，跳过写入")
            return

        # ── 2. Read all indicator data from source via DataReader ──
        base_date_list = []
        if indicator_col_map:
            indicator_dfs = {}
            for aa_code in indicator_col_map:
                ind_data = reader.read_indicator_data(source_sheet, aa_code)
                df = ind_data["data"].copy()

                if df.empty or '日期' not in df.columns:
                    indicator_dfs[aa_code] = pd.DataFrame(columns=['日期'])
                    print(f"    - 源表[{source_sheet}] 未找到指标 {aa_code}，该列跳过")
                    continue

                df['日期'] = pd.to_datetime(df['日期'], errors='coerce')
                df = df[df['日期'].notna() & (df['日期'] >= start_date_ts)].copy()
                df = df.sort_values('日期', ascending=True).reset_index(drop=True)
                indicator_dfs[aa_code] = df

            # ── 3. Build union of all dates across all indicators ──
            all_dates = set()
            for aa_code in indicator_col_map:
                df = indicator_dfs.get(aa_code)
                if df is not None and not df.empty:
                    all_dates.update(df['日期'].tolist())

            if all_dates:
                base_date_list = sorted(all_dates)
            else:
                print(f"    - {source_sheet} 所有指标均无有效数据，仅写入公式")

            # ── 4. Write dates ──
            for i, current_date in enumerate(base_date_list):
                row = start_row + i
                cell = ws.cell(row=row, column=date_col_num, value=current_date)
                cell.number_format = date_format

            # ── 5. Write indicator values aligned by date ──
            for aa_code, col_letter in indicator_col_map.items():
                df = indicator_dfs.get(aa_code)
                if df is None or df.empty:
                    continue

                value_col = df.columns[-1]
                value_map = dict(zip(df['日期'], df[value_col]))
                col_num = column_index_from_string(col_letter)

                for i, date in enumerate(base_date_list):
                    value = value_map.get(date)
                    if value is not None and pd.notna(value):
                        ws.cell(row=start_row + i, column=col_num, value=value)

        num_rows = max(len(base_date_list), params.get('num_rows', 0))
        if not indicator_col_map and formulas:
            # 无指标只有公式时，从 target sheet 的日期列推算行数
            if num_rows == 0:
                r = start_row
                while ws.cell(row=r, column=date_col_num).value is not None:
                    r += 1
                num_rows = r - start_row

        # ── 6. Write formulas ──
        for fm in formulas:
            target_col = fm['col']
            template = fm['template']
            target_col_num = column_index_from_string(target_col)
            fm_start = fm.get('start_row', start_row + 1)

            for i in range(num_rows):
                row = start_row + i
                if row < fm_start:
                    continue
                expr = DecorationPlugin._render_template(template, row)
                ws.cell(row=row, column=target_col_num, value=expr)

        # ── 7. Clear stale data below the written range ──
        last_data_row = start_row + len(base_date_list) - 1
        clear_start = last_data_row + 1
        clear_end = ws.max_row

        cols_to_clear = [date_col_num]
        for col_letter in indicator_col_map.values():
            try:
                cols_to_clear.append(column_index_from_string(col_letter))
            except Exception:
                pass
        for fm in formulas:
            try:
                cols_to_clear.append(column_index_from_string(fm['col']))
            except Exception:
                pass

        for col_num in cols_to_clear:
            for r in range(clear_start, clear_end + 1):
                ws.cell(row=r, column=col_num).value = None

        print(f"    - 完成建筑装饰数据写入: {len(indicator_col_map)} 个指标, "
              f"{len(base_date_list)} 行 (起始行 {start_row})")
        print(f"    - 已清除第 {clear_start}-{clear_end} 行残留数据")
