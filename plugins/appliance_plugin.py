import pandas as pd
from openpyxl.utils import column_index_from_string as col_letter_to_num


class AppliancePlugin:
    @staticmethod
    def appliance_write_sheet(context, params):
        """将源数据按 AA 代码读取后写入目标工作表，每个 section 有独立的日期列。"""
        ws = context['ws']
        reader = context['data_reader']
        sheet_config = context['sheet_config']
        source_sheet = sheet_config['source_sheet']

        sections = sheet_config.get('sections', [])

        # Pre-scan: unmerge all merged cells that overlap with write area
        for section in sections:
            date_col = section['date_col']
            date_col_num = col_letter_to_num(date_col)
            data_start_row = section.get('data_start_row', 4)
            indicators = section['indicators']
            if isinstance(indicators, dict):
                indicators = list(indicators.items())

            # Determine the write range and unmerge
            min_col = date_col_num
            max_col = date_col_num
            if indicators is None or len(indicators) == 0:
                continue
            for _, target_col in indicators:
                c = col_letter_to_num(target_col)
                min_col = min(min_col, c)
                max_col = max(max_col, c)
            # Estimate max data rows (~70)
            max_data_row = data_start_row + 75

            to_unmerge = []
            for mc in ws.merged_cells.ranges:
                if (mc.min_row <= max_data_row and mc.max_row >= data_start_row
                    and mc.min_col <= max_col and mc.max_col >= min_col):
                    to_unmerge.append(str(mc))
            for mc_str in to_unmerge:
                ws.unmerge_cells(mc_str)
            if to_unmerge:
                print(f"    - 取消合并 {len(to_unmerge)} 个单元格区域")

        for section in sections:
            date_col = section['date_col']
            date_col_num = col_letter_to_num(date_col)
            data_start_row = section.get('data_start_row', 4)
            indicators = section.get('indicators') or {}
            if isinstance(indicators, list):
                # Empty list means no indicators, skip
                if not indicators:
                    continue
                # Convert list of dicts/codes to (code, col) pairs
                pairs = []
                for item in indicators:
                    if isinstance(item, dict):
                        pairs.extend(item.items())
                    else:
                        pairs.append((item[0], item[1]) if isinstance(item, (list, tuple)) else (item, None))
                indicators = pairs
            elif isinstance(indicators, dict):
                indicators = list(indicators.items())
            else:
                continue

            # 收集所有指标数据
            all_dates = set()
            code_dfs = {}
            for aa_code, _ in indicators:
                ind_data = reader.read_indicator_data(source_sheet, aa_code)
                df = ind_data["data"]
                if df.empty:
                    continue
                value_col = df.columns[-1]
                df = df.dropna(subset=['日期']).copy()
                if df.empty:
                    continue
                df = df.sort_values('日期', ascending=True).reset_index(drop=True)
                code_dfs[aa_code] = (df, value_col)
                # Include ALL date rows to ensure consistent alignment
                all_dates.update(df['日期'])

            if not code_dfs:
                continue

            date_list = sorted(all_dates, reverse=True)

            # 写入日期列
            for i, date_val in enumerate(date_list):
                row = data_start_row + i
                cell = ws.cell(row=row, column=date_col_num)
                cell.value = date_val.to_pydatetime() if isinstance(date_val, pd.Timestamp) else date_val
                cell.number_format = 'yyyy-mm-dd'

            # 写入各指标值
            if indicators is None or len(indicators) == 0:
                continue
            for aa_code, target_col in indicators:
                pair = code_dfs.get(aa_code)
                if pair is None:
                    continue
                df, value_col = pair
                value_map = dict(zip(df['日期'], df[value_col]))
                target_col_num = col_letter_to_num(target_col)
                for i, date_val in enumerate(date_list):
                    row = data_start_row + i
                    val = value_map.get(date_val)
                    if val is not None and pd.notna(val):
                        cell = ws.cell(row=row, column=target_col_num)
                        cell.value = float(val)
                        cell.number_format = '0.00'

            print(f"    - 完成 {source_sheet} section [{date_col}] 共 {len(indicators)} 个指标")
