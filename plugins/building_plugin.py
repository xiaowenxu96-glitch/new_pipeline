import pandas as pd
import re
import openpyxl
from openpyxl.utils import column_index_from_string as column_letter_to_number, get_column_letter as column_number_to_letter
from datetime import datetime

from core_engine.data_processor import DataProcessor

class BuildingPlugin:
    
    @staticmethod
    def building_write_data(context, params):
        """
        功能说明：
            处理建筑基础指标数据，写入到目标工作表中。
            从 data_reader 缓存中读取指标数据，以第一个指标的日期为基准，
            将所有指标数据按此日期对齐写入 Excel。
            日期按正序排列（最早日期在前），使用日度数据格式。

        params:
            context: dict — Pipeline 上下文，包含 ws、data_reader、sheet_config。
            params: dict — 从 YAML action 配置中解析的参数，包含：
                - start_row: int (默认 11) — 数据写入的起始行号。
                - start_date: str (默认 '2014-01-01') — 数据最早日期。
                - date_format: str (默认 'yyyy-mm-dd') — 日期列的数字格式。
                - indicators: dict — 指标编码到列号的映射（从 action 配置中读取）。
        """
        ws = context['ws']
        reader = context['data_reader']
        sheet_config = context['sheet_config']
        source_sheet = sheet_config['source_sheet']
        indicator_col_map = params.get('indicators') or {}
        start_row = params.get('start_row')
        start_date = params.get('start_date', '2014-01-01')
        date_format = params.get('date_format', 'yyyy-mm-dd')
        start_date_ts = pd.to_datetime(start_date)
        date_columns = params.get('date_columns', None)
        start_column = params.get('start_column', 1)
        unit_conversion = params.get('unit_conversion') or {}

        # 1. 从缓存中获取所需所有指标数据
        indicator_dfs = {}
        for indicator in indicator_col_map:
            ind_data = reader.read_indicator_data(source_sheet, indicator)
            df = ind_data["data"].copy()

            if df.empty or '日期' not in df.columns:
                indicator_dfs[indicator] = pd.DataFrame(columns=['日期'])
                print(f"    - 工作表[{source_sheet}] 未找到指标 {indicator}，该列跳过写入")
                continue

            df['日期'] = pd.to_datetime(df['日期'], errors='coerce')

            # 只保留 start_date 及以后数据
            df = df[df['日期'].notna() & (df['日期'] >= start_date_ts)].copy()
            # 日期按正序排列（最早日期在前）
            df = df.sort_values('日期', ascending=True).reset_index(drop=True)

            indicator_dfs[indicator] = df

        # 2. 使用第一个指标的数据作为基础日期表
        first_indicator = list(indicator_col_map.keys())[0]
        first_df = indicator_dfs[first_indicator].copy()

        if first_df.empty:
            print(f"    - {source_sheet} 的第一个指标 {first_indicator} 无有效数据，跳过写入")
            return

        base_df = first_df.copy()

        # 3. 写入统一日期列和指标数据（应用单位转换）
        # 确定需要写入日期的列列表
        if date_columns is not None:
            # 如果配置了 date_columns，转换为列表格式
            if isinstance(date_columns, int):
                date_cols_list = [date_columns]
            elif isinstance(date_columns, str):
                # 如果是字符串形式的列字母，转换为数字
                try:
                    date_cols_list = [column_letter_to_number(date_columns)]
                except:
                    date_cols_list = [int(date_columns)]
            else:
                # 已经是列表，处理每个元素
                date_cols_list = []
                for col in date_columns:
                    if isinstance(col, str):
                        try:
                            date_cols_list.append(column_letter_to_number(col))
                        except:
                            date_cols_list.append(int(col))
                    else:
                        date_cols_list.append(int(col))
        else:
            # 如果没有配置 date_columns，使用默认的 start_column
            date_cols_list = [start_column]
        
        for i, row in base_df.iterrows():
            current_row = start_row + i
            current_date = row['日期']

            # 在所有配置的日期列中写入日期
            for col_num in date_cols_list:
                if col_num >= 1:
                    date_cell = ws.cell(row=current_row, column=col_num, value=current_date)
                    date_cell.number_format = date_format
        
        
        # 4. 写入指标数据，全部按 base_df 的日期对齐
        base_date_list = base_df['日期'].tolist()

        for indicator, col in list(indicator_col_map.items()):
            df = indicator_dfs[indicator].copy()
            value_col = df.columns[-1]
            value_map = dict(zip(df['日期'], df[value_col]))

            try:
                col_num = column_letter_to_number(col) if isinstance(col, str) else int(col)
                if col_num < 1:
                    print(f"    - 跳过列 {col}，计算出的列号 {col_num} 无效")
                    continue
            except Exception as e:
                print(f"    - 解析列号失败，跳过列 {col}: {e}")
                continue

            # 获取该列的单位转换因子（列字母 → 乘数）
            conversion_factor = unit_conversion.get(col, None)

            for i, date in enumerate(base_date_list):
                value = value_map.get(date, None)
                # 应用单位转换（仅当 value 非空且配置了转换因子时）
                if value is not None and conversion_factor is not None:
                    value = value * conversion_factor
                ws.cell(
                    row=start_row + i,
                    column=col_num,
                    value=value
                )

        print(f"    - 完成建筑基础数据写入")

        # 5. 清除下方残留旧数据，避免图表坐标轴出现 1900 日期
        last_data_row = start_row + len(base_date_list) - 1
        clear_start = last_data_row + 1
        clear_end = 2000  # 足够覆盖图表引用范围

        # 清除 A 列（日期）残留
        for r in range(clear_start, clear_end + 1):
            ws.cell(row=r, column=1, value=None)

        # 清除各指标列残留
        for col in indicator_col_map.values():
            try:
                col_num = column_letter_to_number(col) if isinstance(col, str) else int(col)
                for r in range(clear_start, clear_end + 1):
                    ws.cell(row=r, column=col_num, value=None)
            except Exception:
                pass

        print(f"    - 已清除第 {clear_start}-{clear_end} 行残留数据")

    @staticmethod
    def building_write_formula(context, params):
        """
        通用公式写入函数，支持自定义公式模板

        参数说明：
        - start_row: 起始行号（可选，不提供时通过 start_date + date_col 自动定位）
        - target_formula: 公式模板字符串，支持 {source_column}, {row}, {tname} 等占位符
        - target_column: 目标列列表（写入公式的列）
        - source_column: 源列列表（公式中引用的列）
        - tname: 可选的工作表名称，用于跨表引用
        - format: 可选的数字格式字符串，如 "0.00%"、"#,##0.00"、"[Red](#,##0.00)" 等
        - date_col: 可选的停止检查列（默认为目标列的前一列，即 target_col_num - 1）
        - start_date: 可选，当 start_row 未指定时，在 date_col 列中查找该日期所在行作为起始行

        """
        ws = context['ws']

        start_row = params.get('start_row', None)
        target_formula_template = params.get('target_formula', '')
        target_columns = params.get('target_columns', [])
        source_columns = params.get('source_columns', [])
        tname = params.get('tname', None)  # 可选的工作表名称
        custom_format = params.get('format', None)  # 可选的自定义格式
        date_col = params.get('date_col', "B")  # 可选：停止检查列
        start_date = params.get('start_date', None)  # 可选：起始日期，用于定位 start_row
        # 检查参数有效性
        if not target_formula_template:
            print("    - 缺少 target_formulas 配置，跳过公式写入")
            return
            
        if not target_columns or not source_columns:
            print("    - 缺少 target_columns 或 source_columns 配置，跳过公式写入")
            return
            
        if len(target_columns) != len(source_columns):
            print(f"    - target_columns 和 source_columns 数量不匹配，跳过公式写入")
            return
        
        # 如果提供了 start_date，则查找起始行
        if start_row is None and start_date is not None:
            date_col_for_search = date_col if date_col else target_columns[0]  # 默认使用第一个目标列
            date_col_num = column_letter_to_number(date_col_for_search) if isinstance(date_col_for_search, str) else int(date_col_for_search)
            
            # 在 date_col 列中查找 start_date
            for row in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=date_col_num).value
                if cell_value == start_date or (isinstance(cell_value, datetime) and cell_value.date() == start_date):
                    start_row = row
                    print(f"    - 在列 {date_col_for_search} 中找到起始日期 {start_date}，起始行为 {start_row}")
                    break
            
            if start_row is None:
                print(f"    - 未在列 {date_col_for_search} 中找到起始日期 {start_date}，跳过公式写入")
                return
        
        # 如果 start_row 仍未指定，默认从第2行开始
        if start_row is None:
            start_row = 2
            print(f"    - 未指定起始行，默认从第 {start_row} 行开始")
        
        # 遍历每一对源列和目标列
        for target_col, source_col in zip(target_columns, source_columns):
            try:
                # 转换列标识为数字
                target_col_num = column_letter_to_number(target_col) if isinstance(target_col, str) else int(target_col)
                source_col_num = column_letter_to_number(source_col) if isinstance(source_col, str) else int(source_col)
                
                if target_col_num < 1 or source_col_num < 1:
                    print(f"    - 跳过无效的列配置: target={target_col}, source={source_col}")
                    continue
                
                # 确定停止检查列：优先使用 date_col，否则默认为目标列的前一列
                if date_col is not None:
                    stop_check_col = column_letter_to_number(date_col) if isinstance(date_col, str) else int(date_col)
                    # print(f"    - 使用指定的停止检查列: {date_col}")
                else:
                    stop_check_col = target_col_num - 1
                    if stop_check_col < 1:
                        print(f"    - 警告: 目标列 {target_col} 为第1列，无法检查前一列，将检查目标列自身")
                        stop_check_col = target_col_num
                
                # 从start_row开始，逐行写入公式
                current_row = start_row
                while True:
                    # 使用确定的停止检查列进行判断
                    check_cell = ws.cell(row=current_row, column=stop_check_col)
                    if check_cell.value is None:
                        break
                    
                    # 构建公式：替换模板中的占位符
                    formula = target_formula_template
                    
                    # 先处理 {source_column+N} 和 {source_column-N} 列偏移
                    def replace_source_col_offset(match):
                        offset = int(match.group(1))
                        new_col_num = source_col_num + offset  # 基于源列进行偏移
                        if new_col_num < 1:
                            print(f"警告: 列偏移后列号小于1: {new_col_num}")
                            return match.group(0)  # 返回原样
                        return column_number_to_letter(new_col_num)
                    
                    # 替换 {source_column+N} 和 {source_column-N}
                    formula = re.sub(r'\{source_column([+-]\d+)\}', replace_source_col_offset, formula)
                    
                    # 替换 {source_column} 为实际的源列字母（不带偏移的）
                    source_col_letter = column_number_to_letter(source_col_num)
                    formula = formula.replace('{source_column}', source_col_letter)
                    
                    # 处理行偏移 {row-N} 和 {row+N}
                    formula = re.sub(r'\{row-(\d+)\}', lambda m: str(current_row - int(m.group(1))), formula)
                    formula = re.sub(r'\{row\+(\d+)\}', lambda m: str(current_row + int(m.group(1))), formula)
                    
                    # 替换 {row} 为当前行号（需要放在行偏移之后，避免误匹配）
                    formula = formula.replace('{row}', str(current_row))
                    
                    # 如果有 tname，替换 {tname}
                    if tname:
                        formula = formula.replace('{tname}', tname)
                    
                    # 在目标单元格写入公式
                    target_cell = ws.cell(row=current_row, column=target_col_num)
                    target_cell.value = formula

                    # 设置单元格格式
                    if custom_format:
                        # 如果提供了自定义格式，直接使用
                        target_cell.number_format = custom_format
                    else:
                        # 否则从源列继承格式
                        source_cell = ws.cell(row=current_row, column=source_col_num)
                        target_cell.number_format = source_cell.number_format
                        
                        # 如果源单元格没有特定格式，则使用会计格式（负数用括号表示）
                        if not source_cell.number_format or source_cell.number_format == 'General':
                            target_cell.number_format = '#,##0.00_);(#,##0.00)'

                    current_row += 1
                
                print(f"    - 完成列 {target_col} 的公式写入（共 {current_row - start_row} 行）")
                
            except Exception as e:
                print(f"    - 处理列 {target_col} 时出错: {e}")
                continue

        return

    @staticmethod
    def building_write_formula_with_year_lookup(context, params):
        """
        带年份查找的公式写入函数，自动遍历所有年份列
        
        参数说明：
        - start_row: 起始行号（数据开始行）
        - year_row: 年份所在的行号
        - date_col: 日期列的列字母或数字（存放最新一年的日期）
        - latest_year_col: 最新年份所在的列字母或数字
        - source_column: 数据源列（公式引用的列）
        - source_date_col: 源数据日期列
        - target_formula: 公式模板字符串，支持占位符：
            {source_column} - 源列字母
            {source_column+N} - 源列偏移N列
            {source_date_col} - 源日期列
            {year_col} - 当前年份列字母（自动从年份行读取）
            {year_row} - 年份行号
            {row} - 当前行号
            {row+N} - 当前行偏移N行
            {row-N} - 当前行偏移N行
            {tname} - 可选的工作表名称
        - tname: 可选的工作表名称，用于跨表引用
        - format: 可选的数字格式字符串
        - stop_check_col: 用于判断停止的列（默认为date_col）
        """
        ws = context['ws']
        
        # 读取参数
        start_row = params.get('start_row', 2)
        year_row = params.get('year_row', 1)
        date_col = params.get('date_col', 'A')
        latest_year_col = params.get('latest_year_col', 'Z')
        source_column = params.get('source_column', 'B')
        source_date_col = params.get('source_date_col', 'C')
        target_formula_template = params.get('target_formula', '')
        tname = params.get('tname', None)
        custom_format = params.get('format', None)
        stop_check_col = params.get('stop_check_col', date_col)  # 可选，默认使用date_col
        
        # 参数校验
        if not target_formula_template:
            print("    - 缺少 target_formula 配置，跳过公式写入")
            return
        
        # 转换列标识为数字
        date_col_num = column_letter_to_number(date_col) if isinstance(date_col, str) else int(date_col)
        latest_year_col_num = column_letter_to_number(latest_year_col) if isinstance(latest_year_col, str) else int(latest_year_col)
        source_col_num = column_letter_to_number(source_column) if isinstance(source_column, str) else int(source_column)
        source_date_col_num = column_letter_to_number(source_date_col) if isinstance(source_date_col, str) else int(source_date_col)
        
        # 如果指定了stop_check_col，转换它
        if stop_check_col:
            stop_check_col_num = column_letter_to_number(stop_check_col) if isinstance(stop_check_col, str) else int(stop_check_col)
        else:
            stop_check_col_num = date_col
        
        # 获取所有年份列：从date_col+1到latest_year_col
        year_columns = []
        year_mapping = {}
        for col_num in range(date_col_num + 1, latest_year_col_num + 1):
            year_cell = ws.cell(row=year_row, column=col_num)
            if year_cell.value:
                year_str = str(year_cell.value)
                year_num = re.search(r'(\d{4})', year_str)
                if year_num:
                    year_val = int(year_num.group(1))
                    year_columns.append(col_num)
                    year_mapping[col_num] = year_val
        
        if not year_columns:
            print("    - 警告: 未找到任何年份列，请检查 year_row 和日期列配置")
            return
        
        print(f"    - 找到 {len(year_columns)} 个年份列: {year_mapping}")
        
        # 遍历所有年份列作为目标列
        for target_col_num in year_columns:
            try:
                target_col_letter = column_number_to_letter(target_col_num)
                
                # 从start_row开始，逐行处理
                current_row = start_row
                processed_count = 0
                
                while True:
                    # 检查停止列是否有值
                    check_cell = ws.cell(row=current_row, column=stop_check_col_num)
                    if check_cell.value is None:
                        break
                    
                    # 获取当前行的日期（从源日期列读取，而非 date_col）
                    date_cell = ws.cell(row=current_row, column=source_date_col_num)
                    current_date = date_cell.value
                    
                    if not current_date:
                        current_row += 1
                        continue
                    
                    # 从日期中提取年份
                    if hasattr(current_date, 'year'):
                        current_year = current_date.year
                    else:
                        # 如果是字符串格式的日期
                        date_str = str(current_date)
                        year_match = re.search(r'(\d{4})', date_str)
                        if year_match:
                            current_year = int(year_match.group(1))
                        else:
                            current_row += 1
                            continue
                    
                    # 根据当前行的年份，找到对应的年份列
                    matched_year_col_num = None
                    for col_num, year_val in year_mapping.items():
                        if year_val == current_year:
                            matched_year_col_num = col_num
                            break
                    
                    if matched_year_col_num is None:
                        # 如果当前行的年份没有对应的列，跳过这一行
                        current_row += 1
                        continue
                    
                    # 构建公式
                    formula = target_formula_template
                    
                    # 1. 处理 {source_column+N} 和 {source_column-N} 列偏移
                    def replace_source_col_offset(match):
                        offset = int(match.group(1))
                        new_col_num = source_col_num + offset
                        if new_col_num < 1:
                            print(f"    - 警告: 列偏移后列号小于1: {new_col_num}")
                            return match.group(0)
                        return column_number_to_letter(new_col_num)
                    
                    formula = re.sub(r'\{source_column([+-]\d+)\}', replace_source_col_offset, formula)
                    
                    # 2. 替换 {source_column} 为源列字母
                    source_col_letter = column_number_to_letter(source_col_num)
                    formula = formula.replace('{source_column}', source_col_letter)
                    
                    # 3. 替换 {source_date_col}
                    source_date_col_letter = column_number_to_letter(source_date_col_num)
                    formula = formula.replace('{source_date_col}', source_date_col_letter)

                    # 3.3 替换 {source_col_index} - VLOOKUP中source_column在日期~源列范围内的列号
                    source_col_index = source_col_num - source_date_col_num + 1
                    formula = formula.replace('{source_col_index}', str(source_col_index))

                    # 3.5 替换 {date_col} - 日期列（如K列，存储用于提取月/日的日期）
                    date_col_letter = column_number_to_letter(date_col_num)
                    formula = formula.replace('{date_col}', date_col_letter)

                    # 4. 替换 {year_col} - 当前目标列的字母
                    year_col_letter = column_number_to_letter(target_col_num)
                    formula = formula.replace('{year_col}', year_col_letter)
                    
                    # 5. 替换 {year_row}
                    formula = formula.replace('{year_row}', str(year_row))

                    # 5.5 替换 {year_value} - 当前列对应的纯数字年份（如 2011），从year_mapping中获取
                    year_val = year_mapping[target_col_num]
                    formula = formula.replace('{year_value}', str(year_val))
                    
                    # 6. 处理行偏移 {row-N} 和 {row+N}
                    formula = re.sub(r'\{row-(\d+)\}', lambda m: str(current_row - int(m.group(1))), formula)
                    formula = re.sub(r'\{row\+(\d+)\}', lambda m: str(current_row + int(m.group(1))), formula)
                    
                    # 7. 替换 {row} 为当前行号
                    formula = formula.replace('{row}', str(current_row))
                    
                    # 8. 如果有 tname，替换 {tname}
                    if tname:
                        formula = formula.replace('{tname}', tname)
                    
                    # 写入公式到目标单元格
                    target_cell = ws.cell(row=current_row, column=target_col_num)
                    target_cell.value = formula
                    
                    # 设置单元格格式
                    if custom_format:
                        target_cell.number_format = custom_format
                    else:
                        # 从源列继承格式
                        source_cell = ws.cell(row=current_row, column=source_col_num)
                        target_cell.number_format = source_cell.number_format
                        if not source_cell.number_format or source_cell.number_format == 'General':
                            target_cell.number_format = '#,##0.00_);(#,##0.00)'
                    
                    current_row += 1
                    processed_count += 1
                
                print(f"    - 完成列 {target_col_letter}（{year_mapping[target_col_num]}年）的公式写入（共 {processed_count} 行）")
                
            except Exception as e:
                print(f"    - 处理列 {target_col_letter} 时出错: {e}")
                continue
        
        print(f"    - 所有年份列公式写入完成，共处理 {len(year_columns)} 列")
        return

    @staticmethod
    def building_calc_monthly_from_cumulative(context, params):
        """
        将累计值（月度累计）转换为当月值。支持批量处理多个列对。
        
        参数：
        ----------
        date_column : 日期列（默认"B"）
        source_columns : 累计值列列表（必填）
        target_columns : 输出列列表（必填）
        start_date : 从指定日期开始计算（可选）
        start_row : 从指定行开始计算（可选）
        
        计算逻辑：按月判断，当月份变化时，当月值 = 累计值（重新开始）
        
        示例配置：
        - type: building_calc_monthly_from_cumulative
        date_column: "B"
        source_columns: ["C", "D", "E"]
        target_columns: ["Q", "R", "S"]
        start_date: 2020-01-01
        """
        
        ws = context["ws"]
        
        date_column = params.get("date_column", "B")
        source_columns = params["source_columns"]
        target_columns = params["target_columns"]
        start_date = params.get("start_date", None)
        start_row = params.get("start_row", None)
        
        # 验证列数是否匹配
        if len(source_columns) != len(target_columns):
            print(f"    - 错误：source_columns({len(source_columns)}) 和 target_columns({len(target_columns)}) 数量不匹配")
            return
        
        # 转换日期列号
        date_col = (
            column_letter_to_number(date_column)
            if isinstance(date_column, str)
            else int(date_column)
        )
        
        # 转换所有源列和目标列号
        source_cols = [column_letter_to_number(c) if isinstance(c, str) else int(c) for c in source_columns]
        target_cols = [column_letter_to_number(c) if isinstance(c, str) else int(c) for c in target_columns]
        
        # 确定起始行
        if start_row is None:
            if start_date is None:
                start_row = 2
            else:
                target_date = pd.to_datetime(start_date).date()
                
                start_row = None
                for r in range(1, ws.max_row + 1):
                    value = ws.cell(r, date_col).value
                    
                    if value is None:
                        continue
                    
                    if isinstance(value, datetime):
                        cell_date = value.date()
                    else:
                        try:
                            cell_date = pd.to_datetime(value).date()
                        except Exception:
                            continue
                    
                    if cell_date >= target_date:
                        start_row = r
                        break
                
                if start_row is None:
                    print(f"    - 未找到开始日期 {start_date}")
                    return
        
        # ------------------------
        # 对每组列对进行计算
        # ------------------------
        for idx, (source_col, target_col) in enumerate(zip(source_cols, target_cols)):
            previous_value = None
            previous_year = None
            previous_month = None
            
            row = start_row
            is_first_row = True
            
            while True:
                date_value = ws.cell(row, date_col).value
                
                if date_value is None:
                    break
                
                cumulative = ws.cell(row, source_col).value
                
                try:
                    current_date = pd.to_datetime(date_value)
                except Exception:
                    row += 1
                    continue
                
                if cumulative is None:
                    previous_value = None
                    previous_year = None
                    previous_month = None
                    row += 1
                    continue
                
                try:
                    cumulative = float(cumulative)
                except Exception:
                    previous_value = None
                    previous_month = None
                    row += 1
                    continue
                
                current_year = current_date.strftime("%Y")
                current_month = current_date.strftime("%m")
                
                if is_first_row or previous_month is None or current_year != previous_year:
                    monthly = cumulative
                    is_first_row = False
                else:
                    monthly = cumulative - previous_value
                
                target_cell = ws.cell(row=row, column=target_col)
                target_cell.value = monthly
                target_cell.number_format = ws.cell(row=row, column=source_col).number_format
                
                previous_value = cumulative
                previous_month = current_month
                previous_year = current_year
                
                row += 1

            processed_count = row - start_row
            print(f"    - 完成列 {target_columns[idx]} 的公式写入（共 {processed_count} 行）")

    @staticmethod
    def update_latest_value(context, params):
        """
        通用最新值引用函数。

        参数说明：
        - date_column: str/int — 日期列，用于查找最新日期所在行。
        - target_row: int — 公式写入的目标行号。
        - start_row: int — 从该行开始搜索最新日期。
        - start_column: str/int (默认 "C") — 写入公式的起始列。
        - end_column: str/int (默认 "FA") — 写入公式的结束列。
        - formula_template: str (默认 "={col}{latest_row}") — 公式模板，支持以下占位符：
            {col}       — 当前列的字母（如 C、D、E...）
            {latest_row} — 最新日期所在的行号
            {target_row} — 目标行号（等于 target_row 参数值）
            {row}       — 同 {target_row}

        示例：
            - 默认行为：在 C-FA 列写入 =C{latest_row}、=D{latest_row}...
            - 自定义公式：formula_template: "=IF({col}{latest_row}=\"\",NA(),{col}{latest_row})"
        """
        ws = context["ws"]

        date_column = params.get("date_column", "B")
        target_row = params.get("target_row", None)
        start_row = params.get("start_row", 2)
        start_column = params.get("start_column", "C")
        end_column = params.get("end_column", "FA")
        formula_template = params.get("formula_template", "={col}{latest_row}")

        # target_row 未指定时，默认使用 start_row - 1（数据区域上方一行）
        if target_row is None:
            target_row = start_row - 1

        # 转换日期列号
        date_col = (
            column_letter_to_number(date_column)
            if isinstance(date_column, str)
            else int(date_column)
        )

        # 转换起止列号
        start_col_num = (
            column_letter_to_number(start_column)
            if isinstance(start_column, str)
            else int(start_column)
        )
        end_col_num = (
            column_letter_to_number(end_column)
            if isinstance(end_column, str)
            else int(end_column)
        )

        # 从 start_row 开始，寻找 date_column 列中最大的日期，并得到 latest_row
        max_date = None
        latest_row = None

        for r in range(start_row, ws.max_row + 1):
            date_value = ws.cell(r, date_col).value

            if date_value is None:
                continue

            try:
                if isinstance(date_value, datetime):
                    current_date = date_value.date() if hasattr(date_value, 'date') else date_value
                else:
                    current_date = pd.to_datetime(date_value).date()
                # 排除 NaT（Not a Time）—— pd.NaT 是 datetime 的子类，会绕过上述判断
                if pd.isna(current_date):
                    continue
            except Exception:
                continue

            if max_date is None or current_date > max_date:
                max_date = current_date
                latest_row = r

        if latest_row is None:
            print("    - 未找到有效日期数据，跳过 update_latest_value")
            return

        # 在 target_row 中的 start_column ~ end_column，按 formula_template 写入公式
        for col in range(start_col_num, end_col_num + 1):
            col_letter = column_number_to_letter(col)
            formula = formula_template
            formula = formula.replace("{col}", col_letter)
            # 处理 {latest_row+N} 和 {latest_row-N} 行偏移（放在 {latest_row} 之前）
            formula = re.sub(r'\{latest_row-(\d+)\}', lambda m: str(latest_row - int(m.group(1))), formula)
            formula = re.sub(r'\{latest_row\+(\d+)\}', lambda m: str(latest_row + int(m.group(1))), formula)
            formula = formula.replace("{latest_row}", str(latest_row))
            formula = formula.replace("{target_row}", str(target_row))
            formula = formula.replace("{row}", str(target_row))
            ws.cell(row=target_row, column=col).value = formula

        start_col_letter = column_number_to_letter(start_col_num)
        end_col_letter = column_number_to_letter(end_col_num)
        print(f"    - 已在第{target_row}行 {start_col_letter}-{end_col_letter}列写入公式，引用最新数据行: 第{latest_row}行（日期: {max_date}）")





















