import pandas as pd
from openpyxl.utils import column_index_from_string as col_letter_to_num


class ChemicalPlugin:
    """化工 Pipeline 插件：按 AA 指标代码读取源数据并写入目标文档"""

    @staticmethod
    def _find_indicator_date_column(ws, start_row, indicator_col_num):
        """向左扫描 header 行，找到最近的'日期'列号"""
        for c in range(indicator_col_num - 1, 0, -1):
            v = ws.cell(row=start_row - 1, column=c).value
            if v and str(v).strip() == '日期':
                return c
        return indicator_col_num - 1  # fallback

    @staticmethod
    def _process_indicators_for_date_col(context, date_col_str, indicators, data_start_row):
        """处理一组指标，写入指定的日期列。"""
        ws = context['ws']
        reader = context['data_reader']
        source_sheet = context['sheet_config']['source_sheet']
        start_row = data_start_row
        date_col_num = col_letter_to_num(date_col_str)

        # 收集所有指标数据
        all_dates = set()
        code_dfs = {}
        for aa_code, target_col in indicators.items():
            ind_data = reader.read_indicator_data(source_sheet, aa_code)
            df = ind_data['data']
            if df.empty:
                continue
            df = df[df['日期'].notna()].copy()
            value_col = df.columns[-1]
            df = df[df[value_col].notna()].copy()
            if df.empty:
                continue
            code_dfs[aa_code] = (df, value_col)
            all_dates.update(df['日期'])

        # 即便没有指标，也把日期列清一下（用户可能后续手动填）
        if not indicators or not code_dfs:
            return 0, 0

        date_list = sorted(all_dates, reverse=True)

        # 写入日期列
        for i, date_val in enumerate(date_list):
            row = start_row + i
            c = ws.cell(row=row, column=date_col_num)
            c.value = date_val.to_pydatetime() if isinstance(date_val, pd.Timestamp) else date_val
            c.number_format = 'yyyy-mm-dd'

        # 写入各指标值
        for aa_code, target_col in indicators.items():
            pair = code_dfs.get(aa_code)
            if pair is None:
                continue
            df, value_col = pair
            value_map = dict(zip(df['日期'], df[value_col]))
            tgt_col_num = col_letter_to_num(target_col)
            for i, date_val in enumerate(date_list):
                val = value_map.get(date_val)
                if val is not None:
                    c = ws.cell(row=start_row + i, column=tgt_col_num)
                    c.value = float(val)
                    c.number_format = '0.00'

        return len(date_list), len(indicators)

    @staticmethod
    def chemical_write_sheet(context, params):
        ws = context['ws']
        reader = context['data_reader']
        sheet_config = context['sheet_config']
        defaults = context['defaults']

        source_sheet = sheet_config['source_sheet']
        start_row = sheet_config.get('data_start_row') or sheet_config.get('start_row') or defaults.get('start_row', 44)

        # ---- sections 模式（推荐：每个 section 独立指定 date_col）----
        sections = sheet_config.get('sections')
        if sections:
            total_rows = 0
            total_indicators = 0
            for section in sections:
                sec_date_col = section['date_col']
                sec_indicators = section.get('indicators') or {}
                sec_start = section.get('data_start_row') or start_row
                rows, n_ind = ChemicalPlugin._process_indicators_for_date_col(
                    context, sec_date_col, sec_indicators, sec_start
                )
                total_rows += rows
                total_indicators += n_ind
                if n_ind > 0:
                    print(f"    - [{source_sheet}] section [{sec_date_col}] 写入 {rows} 行, {n_ind} 个指标")
            print(f"    - [{source_sheet}] 共写入 {len(sections)} sections, {total_rows} 行, {total_indicators} 个指标")
            return

        # ---- 旧格式兼容（单个 date_col + indicators）----
        indicators = sheet_config['indicators']
        date_col = sheet_config.get('date_col')

        # 按日期列分组: date_col_num → {dates, [(tgt_col_num, date_map), ...]}
        groups = {}
        for aa_code, target_col in indicators.items():
            ind_data = reader.read_indicator_data(source_sheet, aa_code)
            df = ind_data['data']
            if df.empty:
                continue
            df = df[df['日期'].notna()].copy()
            value_col = df.columns[-1]
            df = df[df[value_col].notna()].copy()
            if df.empty:
                continue

            tgt_col_num = col_letter_to_num(target_col)
            if date_col:
                date_col_num = col_letter_to_num(date_col)
            else:
                date_col_num = ChemicalPlugin._find_indicator_date_column(ws, start_row, tgt_col_num)

            date_val_map = {}
            for _, row in df.iterrows():
                d = row['日期']
                v = row[value_col]
                if pd.notna(d) and pd.notna(v):
                    date_val_map[d] = float(v)

            if date_col_num not in groups:
                groups[date_col_num] = {'dates': set(), 'indicators': []}
            groups[date_col_num]['dates'].update(date_val_map.keys())
            groups[date_col_num]['indicators'].append((tgt_col_num, date_val_map))

        if not groups:
            print(f"    - [{source_sheet}] 未找到任何数据")
            return

        total_rows = 0
        for date_col_num, group in groups.items():
            date_list = sorted(group['dates'], reverse=True)
            for i, date_val in enumerate(date_list):
                row = start_row + i
                c = ws.cell(row=row, column=date_col_num)
                c.value = date_val.to_pydatetime() if isinstance(date_val, pd.Timestamp) else date_val
                c.number_format = 'yyyy-mm-dd'
            for tgt_col_num, date_map in group['indicators']:
                for i, date_val in enumerate(date_list):
                    val = date_map.get(date_val)
                    if val is not None:
                        c = ws.cell(row=start_row + i, column=tgt_col_num)
                        c.value = val
                        c.number_format = '0.00'
            total_rows += len(date_list)

        print(f"    - [{source_sheet}] 写入 {len(groups)} 组日期列, {total_rows} 行, {len(indicators)} 个指标")